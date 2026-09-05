import json
import os
import sys
import unittest
from datetime import date
from io import BytesIO
from types import SimpleNamespace

from fastapi import HTTPException
from pydantic import ValidationError
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database import _ensure_timetable_version_columns
from daily_exports import build_daily_pdf, build_daily_xlsx, daily_export_data, get_period_times
from models import (Base, AbsenceCase, Professor, ScheduleAdjustment, ScheduleAdjustmentLeg,
                    TimetableImportPreview, TimetableLesson, TimetableVersion, SchoolClosure)
from openpyxl import load_workbook
from rescheduling_service import effective_occurrences, version_for_date
from routes.rescheduling import (ActivateTimetableRequest, UpdateTimetableRequest, activate_import,
                                current_timetable, delete_timetable_version, list_timetable_versions,
                                update_timetable_version, get_effective_timetable)


RANGES = [
    {"effective_from": "2027-01-15", "effective_to": "2027-01-30"},
    {"effective_from": "2027-06-21", "effective_to": "2027-07-10"},
]
USER = SimpleNamespace(username="admin")


class TimetableRangesTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(self.engine)
        self.db = sessionmaker(bind=self.engine)()
        self.teacher = Professor(nom="Range Teacher", actiu=True)
        self.normal = TimetableVersion(
            effective_from=date(2026, 9, 1), effective_to=date(2027, 7, 31),
            class_filename="normal.xls", teacher_filename="teachers.xlsx",
        )
        self.db.add_all([self.teacher, self.normal])
        self.db.commit()

    def tearDown(self):
        self.db.close()
        self.engine.dispose()

    def activate(self, ranges=RANGES):
        payload = {
            "format": "post_exam", "issues": [], "warnings": [], "teachers": [self.teacher.nom],
            "summary": {"lessons": 5},
            "lessons": [{"weekday": weekday, "period": 1, "class_code": "1A", "subject": "中文",
                         "teachers": [self.teacher.nom], "movable": True} for weekday in range(5)],
            "teacher_slots": [{"weekday": weekday, "period": 2, "class_code": "", "subject": "值日",
                               "teacher": self.teacher.nom} for weekday in range(5)]
                + [{"weekday": weekday, "period": 1, "class_code": "1A", "subject": "中文",
                    "teacher": self.teacher.nom} for weekday in range(5)],
        }
        self.db.add(TimetableImportPreview(id="ranges", payload=json.dumps(payload),
                                          class_filename="post.xlsx", teacher_filename="teachers.xlsx"))
        self.db.commit()
        result = activate_import("ranges", ActivateTimetableRequest(effective_ranges=ranges), self.db, USER)
        return self.db.get(TimetableVersion, result["version_id"])

    def test_import_round_trip_boundaries_gaps_and_exports(self):
        post = self.activate(list(reversed(RANGES)))
        self.db.expire_all()
        self.assertEqual(json.loads(post.effective_ranges_json), RANGES)
        self.assertEqual((post.effective_from, post.effective_to), (date(2027, 1, 15), date(2027, 7, 10)))
        row = next(row for row in list_timetable_versions(self.db) if row["id"] == post.id)
        self.assertEqual(row["effective_ranges"], RANGES)
        for value in ("2027-01-15", "2027-01-30", "2027-06-21", "2027-07-10"):
            day = date.fromisoformat(value)
            self.assertEqual(version_for_date(self.db, day).id, post.id)
            self.assertEqual(current_timetable(day, self.db)["effective_ranges"], RANGES)
        for value in ("2027-01-14", "2027-01-31", "2027-03-01", "2027-06-20", "2027-07-11"):
            self.assertEqual(version_for_date(self.db, date.fromisoformat(value)).id, self.normal.id)
        for day in (date(2027, 1, 18), date(2027, 6, 21)):
            rows = effective_occurrences(self.db, day, day)
            self.assertEqual({row["version_id"] for row in rows}, {post.id})
            self.assertEqual({row["source"] for row in rows}, {"base", "teacher_workbook"})
            self.assertEqual(get_effective_timetable(day, db=self.db)["effective_ranges"], RANGES)
            self.db.add(AbsenceCase(professor_id=self.teacher.id, data=day, periods_json="[1]", status="resolved"))
            self.db.commit()
            periods = get_period_times(self.db, day)
            self.assertEqual(len(periods), 5)
            entries = daily_export_data(self.db, day)
            workbook = load_workbook(BytesIO(build_daily_xlsx(entries, periods)))
            self.assertEqual(workbook[self.teacher.nom].max_row, 13)
            self.assertTrue(build_daily_pdf(entries, periods).startswith(b"%PDF"))
        self.assertEqual(len(get_period_times(self.db, date(2027, 3, 1))), 9)
        self.assertEqual(effective_occurrences(self.db, date(2027, 1, 30), date(2027, 1, 30)), [])
        self.db.add(SchoolClosure(data=date(2027, 6, 22)))
        self.db.commit()
        self.assertEqual(effective_occurrences(self.db, date(2027, 6, 22), date(2027, 6, 22)), [])

    def test_overlap_uses_matching_start_then_id(self):
        post = self.activate()
        other = TimetableVersion(effective_from=date(2027, 6, 1), effective_to=date(2027, 7, 20),
                                 class_filename="other.xlsx", teacher_filename="teachers.xlsx")
        self.db.add(other)
        self.db.commit()
        day = date(2027, 6, 21)
        self.assertEqual(version_for_date(self.db, day).id, post.id)
        self.assertEqual({row["version_id"] for row in effective_occurrences(self.db, day, day)}, {post.id})
        other.effective_from = day
        self.db.add(TimetableLesson(version_id=other.id, weekday=0, period=1, class_code="1A",
                                    subject="Other", teachers_json="[]"))
        self.db.commit()
        self.assertEqual(version_for_date(self.db, day).id, other.id)
        self.assertEqual({row["version_id"] for row in effective_occurrences(self.db, day, day)}, {other.id})

    def test_edit_legacy_request_and_validation(self):
        post = self.activate()
        invalid = [[], [RANGES[0], RANGES[0]],
                   [{"effective_from": "2027-01-30", "effective_to": "2027-01-15"}],
                   [{"effective_from": "", "effective_to": "2027-01-15"}],
                   [{"effective_from": "2027-01-15"}]]
        for ranges in invalid:
            with self.subTest(ranges=ranges), self.assertRaises((HTTPException, ValidationError)):
                update_timetable_version(post.id, UpdateTimetableRequest(effective_ranges=ranges), self.db, USER)
        with self.assertRaises(HTTPException):
            update_timetable_version(self.normal.id, UpdateTimetableRequest(effective_ranges=RANGES), self.db, USER)
        with self.assertRaises(HTTPException):
            update_timetable_version(post.id, UpdateTimetableRequest(
                effective_from=post.effective_from, effective_to=post.effective_to), self.db, USER)
        update_timetable_version(post.id, UpdateTimetableRequest(effective_ranges=[RANGES[1]]), self.db, USER)
        self.assertEqual(json.loads(post.effective_ranges_json), [RANGES[1]])
        update_timetable_version(post.id, UpdateTimetableRequest(
            effective_from=date(2027, 6, 22), effective_to=date(2027, 7, 10)), self.db, USER)
        self.assertEqual(post.effective_from, date(2027, 6, 22))

    def test_gap_record_does_not_block_import_or_delete(self):
        self.db.add(AbsenceCase(professor_id=self.teacher.id, data=date(2027, 3, 1), periods_json="[1]"))
        self.db.commit()
        post = self.activate()
        row = next(row for row in list_timetable_versions(self.db) if row["id"] == post.id)
        self.assertFalse(row["locked"])
        delete_timetable_version(post.id, self.db, USER)
        self.assertIsNone(self.db.get(TimetableVersion, post.id))

    def test_added_and_removed_periods_protect_records(self):
        post = self.activate([RANGES[0]])
        self.db.add(AbsenceCase(professor_id=self.teacher.id, data=date(2027, 6, 21), periods_json="[1]"))
        self.db.commit()
        with self.assertRaises(HTTPException):
            update_timetable_version(post.id, UpdateTimetableRequest(effective_ranges=RANGES), self.db, USER)
        self.assertEqual(json.loads(post.effective_ranges_json), [RANGES[0]])
        self.db.query(AbsenceCase).delete()
        self.db.commit()
        update_timetable_version(post.id, UpdateTimetableRequest(effective_ranges=RANGES), self.db, USER)
        self.db.add(AbsenceCase(professor_id=self.teacher.id, data=date(2027, 6, 21), periods_json="[1]"))
        self.db.commit()
        with self.assertRaises(HTTPException):
            update_timetable_version(post.id, UpdateTimetableRequest(effective_ranges=[RANGES[0]]), self.db, USER)
        with self.assertRaises(HTTPException):
            delete_timetable_version(post.id, self.db, USER)

    def test_adjustment_count_is_unique_across_periods_and_protected(self):
        post = self.activate()
        lesson = self.db.query(TimetableLesson).filter_by(version_id=post.id, weekday=0).one()
        adjustment = ScheduleAdjustment(kind="direct_swap", status="reverted", locked=False)
        self.db.add(adjustment)
        self.db.flush()
        self.db.add(ScheduleAdjustmentLeg(adjustment_id=adjustment.id, lesson_id=lesson.id,
            class_code="1A", subject="中文", teachers_json="[]", from_date=date(2027, 1, 18),
            from_period=1, to_date=date(2027, 6, 21), to_period=1))
        self.db.commit()
        row = next(row for row in list_timetable_versions(self.db) if row["id"] == post.id)
        self.assertEqual(row["adjustment_records"], 1)
        with self.assertRaises(HTTPException):
            update_timetable_version(post.id, UpdateTimetableRequest(effective_ranges=[RANGES[0]]), self.db, USER)

    def test_import_protects_record_in_second_period(self):
        self.db.add(AbsenceCase(professor_id=self.teacher.id, data=date(2027, 6, 21), periods_json="[1]"))
        self.db.commit()
        with self.assertRaises(HTTPException):
            self.activate()
        self.assertEqual(self.db.query(TimetableVersion).count(), 1)
        self.assertIsNotNone(self.db.get(TimetableImportPreview, "ranges"))

    def test_legacy_migration_is_repeatable_and_preserves_open_range(self):
        engine = create_engine("sqlite:///:memory:")
        with engine.begin() as connection:
            connection.exec_driver_sql("CREATE TABLE timetable_versions (id INTEGER PRIMARY KEY, effective_from DATE)")
            connection.exec_driver_sql("INSERT INTO timetable_versions VALUES (1, '2027-01-15')")
        _ensure_timetable_version_columns(engine)
        _ensure_timetable_version_columns(engine)
        with engine.connect() as connection:
            row = connection.exec_driver_sql("SELECT effective_from, effective_to, effective_ranges_json FROM timetable_versions").one()
            self.assertEqual(tuple(row), ("2027-01-15", None, None))
        engine.dispose()
        self.normal.effective_to = None
        self.db.commit()
        self.assertEqual(version_for_date(self.db, date(2028, 1, 1)).id, self.normal.id)


if __name__ == "__main__":
    unittest.main()
