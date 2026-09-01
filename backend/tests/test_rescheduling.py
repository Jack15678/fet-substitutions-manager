import json
import os
import re
import sys
import unittest
from copy import deepcopy
from datetime import date, datetime
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from fastapi import HTTPException
from pydantic import ValidationError
from sqlalchemy import create_engine
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import sessionmaker

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models import (  # noqa: E402
    AbsenceCase,
    Base,
    Curs,
    Professor,
    ProfessorBaixa,
    ScheduleAdjustment,
    ScheduleAdjustmentLeg,
    ScheduleAudit,
    SchoolClosure,
    TimetableImportPreview,
    TimetableLesson,
    TimetableTeacherSlot,
    TimetableVersion,
)
from daily_exports import build_daily_pdf, build_daily_xlsx, daily_export_data, get_period_times, save_period_times, validate_period_times  # noqa: E402
from database import _ensure_absence_case_columns  # noqa: E402
from openpyxl import Workbook as OpenpyxlWorkbook, load_workbook  # noqa: E402
from repositories import CursRepository  # noqa: E402
from rescheduling_service import apply_import_resolutions, absence_keys, analyze_absence, build_import_preview, choose_global, effective_occurrences, get_schedule_revision, version_for_date  # noqa: E402
from routes.rescheduling import (  # noqa: E402
    AbsenceBatchCreateRequest,
    AbsenceCreateRequest,
    ActivateTimetableRequest,
    ClosureInput,
    ClosureListRequest,
    ConfirmRequest,
    ManualCoverRequest,
    ReschedulingConfigRequest,
    UpdateTimetableRequest,
    UpdateAdjustmentRequest,
    SaveImportResolutionsRequest,
    create_absences_batch,
    activate_import,
    confirm_adjustment,
    confirm_manual_cover,
    discard_import_preview,
    delete_adjustment,
    delete_timetable_version,
    export_daily_pdf,
    export_daily_xlsx,
    get_effective_timetable,
    get_closures,
    list_records,
    list_manual_arrangements,
    list_timetable_versions,
    purge_absence,
    replace_closures,
    revert_adjustment,
    save_import_resolutions,
    update_absence,
    update_adjustment,
    update_rescheduling_config,
    update_timetable_version,
    teacher_statistics,
)


class ReschedulingTests(unittest.TestCase):
    def setUp(self):
        engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(engine)
        self.db = sessionmaker(bind=engine)()
        self.now_patcher = patch(
            "routes.rescheduling.hong_kong_now", return_value=datetime(2026, 8, 10, 8, 0)
        )
        self.now_patcher.start()

    def tearDown(self):
        self.now_patcher.stop()
        self.db.close()

    def test_daily_export_without_records_has_clear_error(self):
        for export in (export_daily_xlsx, export_daily_pdf):
            with self.subTest(export=export.__name__), self.assertRaises(HTTPException) as raised:
                export(date(2026, 8, 30), self.db, SimpleNamespace(username="admin"))
            self.assertEqual(raised.exception.status_code, 404)
            self.assertEqual(raised.exception.detail, "所選日期沒有可匯出的調課／代課記錄")

    def test_daily_export_excludes_pending_absences(self):
        teacher = Professor(nom="待安排老師", actiu=True)
        self.db.add(teacher)
        self.db.flush()
        self.db.add(AbsenceCase(
            professor_id=teacher.id, data=date(2026, 8, 30), periods_json="[7, 8, 9]", status="open",
        ))
        self.db.commit()

        self.assertEqual(daily_export_data(self.db, date(2026, 8, 30)), [])
        for export in (export_daily_xlsx, export_daily_pdf):
            with self.subTest(export=export.__name__), self.assertRaises(HTTPException) as raised:
                export(date(2026, 8, 30), self.db, SimpleNamespace(username="admin"))
            self.assertEqual(raised.exception.status_code, 404)

    def test_absence_reason_contract_and_legacy_migration(self):
        with self.assertRaises(ValidationError):
            AbsenceCreateRequest(professor_id=1, data=date(2026, 8, 10), periods=[1])
        with self.assertRaises(ValidationError):
            AbsenceCreateRequest(
                professor_id=1, data=date(2026, 8, 10), periods=[1],
                reason_type="other", reason_detail="x" * 201,
            )
        other = AbsenceCreateRequest(
            professor_id=1, data=date(2026, 8, 10), periods=[1],
            reason_type="other", reason_detail="  可選補充  ",
        )
        preset = AbsenceCreateRequest(
            professor_id=1, data=date(2026, 8, 10), periods=[1],
            reason_type="sick", reason_detail="不應保存",
        )
        blank = AbsenceCreateRequest(
            professor_id=1, data=date(2026, 8, 10), periods=[1],
            reason_type="other", reason_detail="   ",
        )
        for reason_type in ("sick", "follow_up", "team_training", "other"):
            self.assertEqual(AbsenceCreateRequest(
                professor_id=1, data=date(2026, 8, 10), periods=[1], reason_type=reason_type,
            ).reason_type, reason_type)
        self.assertEqual(other.reason_detail, "可選補充")
        self.assertIsNone(preset.reason_detail)
        self.assertIsNone(blank.reason_detail)

        engine = create_engine("sqlite:///:memory:")
        with engine.begin() as connection:
            connection.exec_driver_sql(
                "CREATE TABLE absence_cases (id INTEGER PRIMARY KEY, periods_json TEXT NOT NULL)"
            )
            connection.exec_driver_sql("INSERT INTO absence_cases VALUES (1, '[1]')")
        _ensure_absence_case_columns(engine)
        with engine.connect() as connection:
            columns = {row[1] for row in connection.exec_driver_sql("PRAGMA table_info(absence_cases)")}
            legacy = connection.exec_driver_sql(
                "SELECT reason_type, reason_detail FROM absence_cases WHERE id = 1"
            ).one()
        self.assertTrue({"reason_type", "reason_detail"}.issubset(columns))
        self.assertEqual(legacy, (None, None))
        engine.dispose()

    def test_academic_years_keep_explicit_summer_gap(self):
        first = CursRepository.create(
            self.db, "2025-2026", date(2025, 9, 1), date(2026, 6, 30),
        )
        second = CursRepository.create(
            self.db, "2026-2027", date(2026, 9, 1), date(2027, 6, 30),
        )

        self.assertEqual(first.data_fi, date(2026, 6, 30))
        self.assertIsNone(CursRepository.get_for_date(self.db, "2026-07-15"))
        self.assertEqual(CursRepository.get_for_date(self.db, "2026-09-01").id, second.id)
        with self.assertRaises(ValueError):
            CursRepository.create(
                self.db, "重疊", date(2026, 6, 1), date(2026, 8, 31),
            )

    def test_calendar_year_holiday_replace_keeps_other_years(self):
        self.db.add_all([
            SchoolClosure(data=date(2025, 12, 25)),
            SchoolClosure(data=date(2026, 12, 25)),
        ])
        self.db.commit()

        replace_closures(
            ClosureListRequest(year=2026, closures=[
                ClosureInput(data=date(2026, 1, 1)),
                ClosureInput(data=date(2026, 1, 1)),
            ]),
            self.db,
            SimpleNamespace(username="admin"),
        )

        self.assertEqual(get_closures(2026, self.db), [{"date": "2026-01-01", "note": None}])
        self.assertEqual(get_closures(2025, self.db), [{"date": "2025-12-25", "note": None}])
        with self.assertRaises(HTTPException):
            replace_closures(
                ClosureListRequest(
                    year=2026,
                    closures=[ClosureInput(data=date(2027, 1, 1))],
                ),
                self.db,
                SimpleNamespace(username="admin"),
            )

    def test_batch_absence_creates_independent_cases_atomically(self):
        teachers = [Professor(nom=f"{name}老師", actiu=True) for name in ("A", "B", "C")]
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), effective_to=date(2026, 8, 31),
            class_filename="classes.xls", teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([*teachers, version])
        self.db.flush()
        self.db.add_all([
            TimetableLesson(version_id=version.id, weekday=0, period=1, class_code="1A", subject="中文", teachers_json=json.dumps([teachers[0].id])),
            TimetableLesson(version_id=version.id, weekday=0, period=2, class_code="2A", subject="英文", teachers_json=json.dumps([teachers[1].id])),
            TimetableLesson(version_id=version.id, weekday=1, period=3, class_code="3A", subject="數學", teachers_json=json.dumps([teachers[2].id])),
        ])
        self.db.commit()

        result = create_absences_batch(
            AbsenceBatchCreateRequest(
                items=[
                    AbsenceCreateRequest(professor_id=teachers[0].id, data=date(2026, 8, 10), periods=[1], reason_type="sick"),
                    AbsenceCreateRequest(professor_id=teachers[1].id, data=date(2026, 8, 10), periods=[2], reason_type="other", reason_detail="  家庭安排  "),
                    AbsenceCreateRequest(professor_id=teachers[2].id, data=date(2026, 8, 11), periods=[3], reason_type="official", reason_detail="ignored"),
                ],
            ),
            self.db,
            SimpleNamespace(username="admin"),
        )

        rows = self.db.query(AbsenceCase).order_by(AbsenceCase.id).all()
        self.assertEqual(len(rows), 3)
        self.assertEqual([json.loads(row.periods_json) for row in rows], [[1], [2], [3]])
        self.assertEqual(
            [(row.reason_type, row.reason_detail) for row in rows],
            [("sick", None), ("other", "家庭安排"), ("official", None)],
        )
        audit = json.loads(self.db.query(ScheduleAudit).order_by(ScheduleAudit.id).first().detail_json)
        self.assertEqual(audit["reason_type"], "sick")
        self.assertEqual(result["batch_absence_case_ids"], [row.id for row in rows])
        self.assertEqual([item["date"] for item in result["analyses"]], ["2026-08-10", "2026-08-11"])

        with self.assertRaises(HTTPException):
            create_absences_batch(
                AbsenceBatchCreateRequest(
                    items=[
                        AbsenceCreateRequest(professor_id=teachers[0].id, data=date(2026, 8, 10), periods=[1], reason_type="sick"),
                        AbsenceCreateRequest(professor_id=999999, data=date(2026, 8, 10), periods=[1], reason_type="sick"),
                    ],
                ),
                self.db,
                SimpleNamespace(username="admin"),
            )
        self.assertEqual(self.db.query(AbsenceCase).count(), 3)

    def test_manual_arrangement_ranks_free_neighbors_and_resolves_absence(self):
        absent, clear, one_busy, two_busy, extra = [
            Professor(nom=f"{name}老師", actiu=True)
            for name in ("缺席", "前後空", "一邊忙", "兩邊忙", "另一位")
        ]
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), effective_to=date(2026, 8, 31),
            class_filename="classes.xls", teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([absent, clear, one_busy, two_busy, extra, version])
        self.db.flush()
        self.db.add_all([
            TimetableLesson(
                version_id=version.id, weekday=0, period=3, class_code="1A",
                subject="中文", teachers_json=json.dumps([absent.id]),
            ),
            TimetableLesson(
                version_id=version.id, weekday=0, period=1, class_code="2A",
                subject="英文", teachers_json=json.dumps([clear.id]),
            ),
            TimetableLesson(
                version_id=version.id, weekday=0, period=2, class_code="3A",
                subject="英文", teachers_json=json.dumps([one_busy.id]),
            ),
            TimetableLesson(
                version_id=version.id, weekday=0, period=2, class_code="4A",
                subject="英文", teachers_json=json.dumps([two_busy.id]),
            ),
            TimetableLesson(
                version_id=version.id, weekday=0, period=4, class_code="4A",
                subject="常識", teachers_json=json.dumps([two_busy.id]),
            ),
            TimetableLesson(
                version_id=version.id, weekday=0, period=2, class_code="5A",
                subject="英文", teachers_json=json.dumps([extra.id]),
            ),
            TimetableLesson(
                version_id=version.id, weekday=0, period=4, class_code="5A",
                subject="常識", teachers_json=json.dumps([extra.id]),
            ),
        ])
        absence = AbsenceCase(
            professor_id=absent.id, data=date(2026, 8, 10), periods_json="[3]",
            status="open", created_by="admin",
        )
        self.db.add(absence)
        self.db.commit()

        queue = list_manual_arrangements(self.db, SimpleNamespace(username="admin"))
        self.assertEqual(len(queue["tasks"]), 1)
        task = queue["tasks"][0]
        self.assertEqual(len(task["candidates"]), 4)
        self.assertEqual(task["candidates"][0]["id"], clear.id)
        self.assertEqual(
            {item["id"] for item in task["candidates"]},
            {clear.id, one_busy.id, two_busy.id, extra.id},
        )
        self.assertEqual(
            [item["adjacent_busy_count"] for item in task["candidates"]],
            [0, 1, 2, 2],
        )

        result = confirm_manual_cover(
            ManualCoverRequest(
                absence_case_id=absence.id,
                occurrence_id=task["target"]["occurrence_id"],
                replacement_teacher_id=clear.id,
                expected_revision=queue["revision"],
            ),
            self.db,
            SimpleNamespace(username="admin"),
        )

        self.db.refresh(absence)
        self.assertEqual(absence.status, "resolved")
        self.assertEqual(result["absence_case_id"], absence.id)
        self.assertEqual(result["kind"], "emergency_cover")
        self.assertEqual(result["legs"][0]["replacement_teacher_id"], clear.id)

    def test_recommended_task_can_use_existing_manual_cover_flow(self):
        absent, swap_teacher, manual_teacher = [
            Professor(nom=name, actiu=True)
            for name in ("缺席老師全名", "互調老師全名", "人工老師全名")
        ]
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), effective_to=date(2026, 8, 31),
            class_filename="classes.xls", teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([absent, swap_teacher, manual_teacher, version])
        self.db.flush()
        self.db.add_all([
            TimetableLesson(
                version_id=version.id, weekday=0, period=1, class_code="1A",
                subject="中文", teachers_json=json.dumps([absent.id]),
            ),
            TimetableLesson(
                version_id=version.id, weekday=0, period=4, class_code="1A",
                subject="英文", teachers_json=json.dumps([swap_teacher.id]),
            ),
            TimetableLesson(
                version_id=version.id, weekday=1, period=1, class_code="2A",
                subject="中文", teachers_json=json.dumps([manual_teacher.id]),
            ),
        ])
        absence = AbsenceCase(
            professor_id=absent.id, data=date(2026, 8, 10), periods_json="[1]",
            reason_type="sick", status="open", created_by="admin",
        )
        self.db.add(absence)
        self.db.commit()

        queue = list_manual_arrangements(self.db, SimpleNamespace(username="admin"))
        task = queue["tasks"][0]
        self.assertEqual(task["status"], "recommended")
        self.assertIn(manual_teacher.id, {item["id"] for item in task["candidates"]})

        result = confirm_manual_cover(
            ManualCoverRequest(
                absence_case_id=absence.id,
                occurrence_id=task["target"]["occurrence_id"],
                replacement_teacher_id=manual_teacher.id,
                expected_revision=queue["revision"],
            ),
            self.db,
            SimpleNamespace(username="admin"),
        )

        self.assertEqual(result["kind"], "emergency_cover")
        self.assertEqual(result["legs"][0]["replacement_teacher_name"], manual_teacher.nom)
        self.db.refresh(absence)
        self.assertEqual(absence.status, "resolved")

    def test_co_taught_lesson_waits_for_manual_confirmation(self):
        absent, co_teacher = [Professor(nom=name, actiu=True) for name in ("缺席老師", "在場老師")]
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), effective_to=date(2026, 8, 31),
            class_filename="classes.xls", teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([absent, co_teacher, version])
        self.db.flush()
        lesson = TimetableLesson(
            version_id=version.id, weekday=0, period=1, class_code="1C",
            subject="圖書", teachers_json=json.dumps([absent.id, co_teacher.id]),
        )
        absence = AbsenceCase(
            professor_id=absent.id, data=date(2026, 8, 10), periods_json="[1]",
            status="open", created_by="admin",
        )
        self.db.add_all([lesson, absence])
        self.db.commit()

        analysis = analyze_absence(self.db, absence)
        self.assertEqual(analysis["tasks"][0]["status"], "unresolved")
        self.assertEqual(analysis["tasks"][0]["alternatives"], [])
        queue = list_manual_arrangements(self.db, SimpleNamespace(username="admin"))
        self.assertEqual(queue["tasks"][0]["co_teachers"], [{"id": co_teacher.id, "name": co_teacher.nom}])

        result = confirm_manual_cover(
            ManualCoverRequest(
                absence_case_id=absence.id,
                occurrence_id=queue["tasks"][0]["target"]["occurrence_id"],
                co_teacher_only=True,
                expected_revision=queue["revision"],
            ),
            self.db,
            SimpleNamespace(username="admin"),
        )

        self.db.refresh(absence)
        effective = effective_occurrences(self.db, absence.data, absence.data)
        self.assertEqual(absence.status, "resolved")
        self.assertEqual(result["kind"], "co_teacher_solo")
        self.assertEqual(effective[0]["teachers"], [co_teacher.id])

    def test_batch_absence_appends_periods_and_preserves_existing_reason(self):
        teacher = Professor(nom="A老師", actiu=True)
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), effective_to=date(2026, 8, 31),
            class_filename="classes.xls", teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([teacher, version])
        self.db.flush()
        self.db.add_all([
            TimetableLesson(
                version_id=version.id, weekday=0, period=period, class_code="1A",
                subject=subject, teachers_json=json.dumps([teacher.id]),
            )
            for period, subject in ((3, "中文"), (4, "英文"))
        ])
        existing = AbsenceCase(
            professor_id=teacher.id, data=date(2026, 8, 10), periods_json="[3]",
            reason_type="sick", created_by="admin",
        )
        self.db.add(existing)
        self.db.commit()

        result = create_absences_batch(
            AbsenceBatchCreateRequest(
                items=[AbsenceCreateRequest(professor_id=teacher.id, data=date(2026, 8, 10), periods=[4], reason_type="training")],
            ),
            self.db,
            SimpleNamespace(username="admin"),
        )

        self.assertEqual(json.loads(existing.periods_json), [3, 4])
        self.assertEqual(existing.reason_type, "sick")
        self.assertEqual(result["updated_absence_case_ids"], [existing.id])
        self.assertEqual(
            [task["target"]["period"] for task in result["tasks"]],
            [3, 4],
        )

    def test_repeated_overlapping_appends_are_monotonic_and_idempotent(self):
        teacher = Professor(nom="連續追加老師", actiu=True)
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), effective_to=date(2026, 8, 31),
            class_filename="classes.xls", teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([teacher, version])
        self.db.flush()
        self.db.add_all([
            TimetableLesson(
                version_id=version.id, weekday=0, period=period, class_code="1A",
                subject=f"課堂{period}", teachers_json=json.dumps([teacher.id]),
            )
            for period in range(1, 10)
        ])
        existing = AbsenceCase(
            professor_id=teacher.id, data=date(2026, 8, 10), periods_json="[1]",
            reason_type="sick", created_by="admin",
        )
        self.db.add(existing)
        self.db.commit()

        for period in range(2, 10):
            result = create_absences_batch(
                AbsenceBatchCreateRequest(items=[AbsenceCreateRequest(
                    professor_id=teacher.id, data=date(2026, 8, 10),
                    periods=[period - 1, period], reason_type="training",
                )]),
                self.db,
                SimpleNamespace(username="admin"),
            )
            self.assertEqual(result["updated_absence_case_ids"], [existing.id])

        self.db.refresh(existing)
        self.assertEqual(json.loads(existing.periods_json), list(range(1, 10)))
        self.assertEqual(existing.reason_type, "sick")
        self.assertEqual(
            self.db.query(AbsenceCase).filter(
                AbsenceCase.professor_id == teacher.id,
                AbsenceCase.data == date(2026, 8, 10),
                AbsenceCase.status != "cancelled",
            ).count(),
            1,
        )
        revision = get_schedule_revision(self.db)
        audit_count = self.db.query(ScheduleAudit).filter_by(
            action="update", entity_type="absence_case", entity_id=str(existing.id),
        ).count()
        self.assertEqual((revision, audit_count), (8, 8))

        unchanged = create_absences_batch(
            AbsenceBatchCreateRequest(items=[AbsenceCreateRequest(
                professor_id=teacher.id, data=date(2026, 8, 10),
                periods=list(range(1, 10)), reason_type="other",
            )]),
            self.db,
            SimpleNamespace(username="admin"),
        )
        self.assertEqual(unchanged["updated_absence_case_ids"], [])
        self.assertEqual(get_schedule_revision(self.db), revision)
        self.assertEqual(
            self.db.query(ScheduleAudit).filter_by(
                action="update", entity_type="absence_case", entity_id=str(existing.id),
            ).count(),
            audit_count,
        )

    def test_mixed_batch_append_and_create_stay_isolated(self):
        existing_teacher = Professor(nom="已有缺席老師", actiu=True)
        new_teacher = Professor(nom="新缺席老師", actiu=True)
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), effective_to=date(2026, 8, 31),
            class_filename="classes.xls", teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([existing_teacher, new_teacher, version])
        self.db.flush()
        self.db.add_all([
            TimetableLesson(
                version_id=version.id, weekday=0, period=period, class_code=class_code,
                subject=subject, teachers_json=json.dumps([teacher_id]),
            )
            for teacher_id, period, class_code, subject in (
                (existing_teacher.id, 1, "1A", "中文"),
                (existing_teacher.id, 3, "1A", "英文"),
                (new_teacher.id, 2, "1B", "數學"),
            )
        ])
        existing = AbsenceCase(
            professor_id=existing_teacher.id, data=date(2026, 8, 10), periods_json="[1]",
            reason_type="sick", created_by="admin",
        )
        self.db.add(existing)
        self.db.commit()

        result = create_absences_batch(
            AbsenceBatchCreateRequest(items=[
                AbsenceCreateRequest(
                    professor_id=existing_teacher.id, data=date(2026, 8, 10),
                    periods=[3], reason_type="training",
                ),
                AbsenceCreateRequest(
                    professor_id=new_teacher.id, data=date(2026, 8, 10),
                    periods=[2], reason_type="other",
                ),
            ]),
            self.db,
            SimpleNamespace(username="admin"),
        )

        new_absence = self.db.query(AbsenceCase).filter_by(professor_id=new_teacher.id).one()
        self.assertEqual(json.loads(existing.periods_json), [1, 3])
        self.assertEqual(existing.reason_type, "sick")
        self.assertEqual(json.loads(new_absence.periods_json), [2])
        self.assertEqual(new_absence.reason_type, "other")
        self.assertEqual(result["updated_absence_case_ids"], [existing.id])
        self.assertEqual(result["created_absence_case_ids"], [new_absence.id])
        self.assertEqual(result["batch_absence_case_ids"], [existing.id, new_absence.id])
        self.assertEqual(self.db.query(AbsenceCase).count(), 2)

    def test_only_one_active_absence_per_teacher_and_date(self):
        teacher = Professor(nom="唯一老師", actiu=True)
        self.db.add(teacher)
        self.db.flush()
        self.db.add(AbsenceCase(
            professor_id=teacher.id, data=date(2026, 8, 10), periods_json="[1]",
        ))
        self.db.commit()

        self.db.add(AbsenceCase(
            professor_id=teacher.id, data=date(2026, 8, 10), periods_json="[2]",
        ))
        with self.assertRaises(IntegrityError):
            self.db.commit()
        self.db.rollback()

        self.db.add(AbsenceCase(
            professor_id=teacher.id, data=date(2026, 8, 10), periods_json="[2]", status="cancelled",
        ))
        self.db.commit()

    @patch("routes.rescheduling.hong_kong_today", return_value=date(2026, 8, 10))
    def test_batch_absence_with_no_lesson_creates_no_record_or_export(self, _today):
        teacher = Professor(nom="A老師", actiu=True)
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), effective_to=date(2026, 8, 31),
            class_filename="classes.xls", teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([teacher, version])
        self.db.flush()
        self.db.add(TimetableLesson(
            version_id=version.id, weekday=0, period=2, class_code="1A",
            subject="中文", teachers_json=json.dumps([teacher.id]),
        ))
        self.db.commit()

        with self.assertRaises(HTTPException):
            create_absences_batch(
                AbsenceBatchCreateRequest(
                    items=[AbsenceCreateRequest(professor_id=teacher.id, data=date(2026, 8, 10), periods=[1], reason_type="sick")],
                ),
                self.db,
                SimpleNamespace(username="admin"),
            )

        self.assertEqual(self.db.query(AbsenceCase).count(), 0)
        self.assertEqual(list_records(scope="today", page=1, page_size=20, db=self.db)["total"], 0)
        self.assertEqual(daily_export_data(self.db, date(2026, 8, 10)), [])

    def test_timetable_version_crud_preserves_linked_records(self):
        teacher = Professor(nom="A老師", actiu=True)
        old_version = TimetableVersion(
            effective_from=date(2026, 8, 10), effective_to=date(2026, 8, 16), class_filename="old.xls",
            teacher_filename="old.xlsx", active=False,
        )
        unused_version = TimetableVersion(
            effective_from=date(2026, 8, 17), effective_to=date(2026, 12, 31), class_filename="new.xls",
            teacher_filename="new.xlsx", active=True,
        )
        self.db.add_all([teacher, old_version, unused_version])
        self.db.flush()
        lesson = TimetableLesson(
            version_id=old_version.id, weekday=0, period=1, class_code="1A",
            subject="中文", teachers_json=json.dumps([teacher.id]),
        )
        absence = AbsenceCase(
            professor_id=teacher.id, data=date(2026, 8, 11), periods_json="[1]",
        )
        adjustment = ScheduleAdjustment(kind="direct_swap", status="reverted", locked=False)
        self.db.add_all([lesson, absence, adjustment])
        self.db.flush()
        self.db.add(ScheduleAdjustmentLeg(
            adjustment_id=adjustment.id, lesson_id=lesson.id, class_code="1A", subject="中文",
            teachers_json=json.dumps([teacher.id]), from_date=date(2026, 8, 11), from_period=1,
            to_date=date(2026, 8, 11), to_period=2,
        ))
        self.db.commit()

        rows = list_timetable_versions(self.db)
        locked = next(row for row in rows if row["id"] == old_version.id)
        self.assertTrue(locked["locked"])
        self.assertEqual((locked["absence_records"], locked["adjustment_records"]), (1, 1))

        user = SimpleNamespace(username="admin")
        update_timetable_version(
            unused_version.id,
            UpdateTimetableRequest(effective_from=date(2026, 8, 18), effective_to=date(2026, 12, 31)),
            self.db, user,
        )
        delete_timetable_version(unused_version.id, self.db, user)
        self.assertIsNone(self.db.get(TimetableVersion, unused_version.id))
        self.assertIsNone(version_for_date(self.db, date(2026, 8, 17)))
        with self.assertRaises(HTTPException) as blocked_delete:
            delete_timetable_version(old_version.id, self.db, user)
        self.assertEqual(blocked_delete.exception.status_code, 409)

    def test_separate_post_exam_timetables_override_normal_range(self):
        periods = [
            ("8:25-9:15", "數學"), ("9:15-10:05", "人文"),
            ("10:25-11:15", "英文"), ("11:15-12:05", "中文"),
            ("12:25-13:00", "科學"),
        ]
        class_workbook = OpenpyxlWorkbook()
        class_sheet = class_workbook.active
        class_sheet.title = "班別試後時間表"
        classes = [
            f"{grade}{letter}"
            for grade, letters in ((1, "ABCD"), (2, "ABCD"), (3, "ABCDE"),
                                   (4, "ABCD"), (5, "ABCD"), (6, "ABCD"))
            for letter in letters
        ]
        class_sheet.append(["班別", *classes])
        class_sheet.append(["班主任", *(["丁依敏"] * len(classes))])
        class_sheet.append(["8:10-8:25", *(["班主任課"] * len(classes))])
        for time, subject in periods[:2]:
            class_sheet.append([time, *([subject] * len(classes))])
            class_sheet.append(["", *(["丁依敏"] * len(classes))])
        class_sheet.append(["10:05-10:25", "小息"])
        for time, subject in periods[2:4]:
            class_sheet.append([time, *([subject] * len(classes))])
            class_sheet.append(["", *(["丁依敏"] * len(classes))])
        class_sheet.append(["12:05-12:25", "小息"])
        class_sheet.append([periods[4][0], *([periods[4][1]] * len(classes))])
        class_sheet.append(["", *(["丁依敏"] * len(classes))])

        teacher_workbook = OpenpyxlWorkbook()
        teacher_sheet = teacher_workbook.active
        teacher_sheet.title = "教師試後時間表"
        teachers = ["丁依敏", *[f"測試教師{index}" for index in range(2, 52)]]
        teacher_sheet.append(["編號", *range(1, 52)])
        teacher_sheet.append(["教師", *teachers])
        teacher_sheet.append(["8:10-8:25", *([""] * 51)])
        for time, subject in periods[:2]:
            teacher_sheet.append([time, subject, *([""] * 50)])
            teacher_sheet.append(["", "、".join(classes), *([""] * 50)])
        teacher_sheet.append(["10:05-10:25", "小息"])
        for time, subject in periods[2:4]:
            teacher_sheet.append([time, subject, *([""] * 50)])
            teacher_sheet.append(["", "、".join(classes), *([""] * 50)])
        teacher_sheet.append(["12:05-12:25", "小息"])
        teacher_sheet.append([periods[4][0], periods[4][1], *([""] * 50)])
        teacher_sheet.append(["", "、".join(classes), *([""] * 50)])

        class_content, teacher_content = BytesIO(), BytesIO()
        class_workbook.save(class_content)
        teacher_workbook.save(teacher_content)
        payload = build_import_preview(
            class_content.getvalue(), teacher_content.getvalue(), post_exam=True
        )
        self.assertEqual(payload["summary"], {
            "classes": 25, "teachers": 51, "lessons": 625,
            "teacher_slots": 625, "blocked_lessons": 0, "issues": 0,
        })

        normal_teacher = Professor(nom="正常老師", actiu=True)
        normal = TimetableVersion(
            effective_from=date(2026, 2, 1), effective_to=date(2026, 7, 15),
            class_filename="normal.xls", teacher_filename="normal.xlsx", active=True,
        )
        self.db.add_all([normal_teacher, normal])
        self.db.flush()
        self.db.add(TimetableLesson(
            version_id=normal.id, weekday=0, period=1, class_code="1A",
            subject="正常課程", teachers_json=json.dumps([normal_teacher.id]),
        ))
        self.db.add(TimetableImportPreview(
            id="post-exam", payload=json.dumps(payload, ensure_ascii=False),
            class_filename="試後班別表.xlsx", teacher_filename="試後教師表.xlsx", created_by="admin",
        ))
        self.db.commit()

        result = activate_import(
            "post-exam",
            ActivateTimetableRequest(
                effective_from=date(2026, 6, 22), effective_to=date(2026, 7, 8),
            ),
            self.db,
            SimpleNamespace(username="admin"),
        )
        self.assertEqual(result["lessons"], 625)
        self.assertEqual(version_for_date(self.db, date(2026, 6, 15)).id, normal.id)
        self.assertNotEqual(version_for_date(self.db, date(2026, 6, 22)).id, normal.id)
        self.assertEqual(version_for_date(self.db, date(2026, 7, 13)).id, normal.id)
        subjects = {row["subject"] for row in effective_occurrences(
            self.db, date(2026, 6, 22), date(2026, 6, 22)
        )}
        self.assertEqual(subjects, {subject for _, subject in periods})

    def test_analysis_never_swaps_lessons_between_timetable_versions(self):
        normal_teacher = Professor(nom="Normal Teacher", actiu=True)
        post_exam_teacher = Professor(nom="Post-exam Teacher", actiu=True)
        normal = TimetableVersion(
            effective_from=date(2026, 2, 1), effective_to=date(2026, 7, 15),
            class_filename="normal.xls", teacher_filename="normal.xlsx", active=True,
        )
        post_exam = TimetableVersion(
            effective_from=date(2026, 6, 22), effective_to=date(2026, 7, 8),
            class_filename="post-exam.xlsx", teacher_filename="post-exam-teacher.xlsx", active=True,
        )
        self.db.add_all([normal_teacher, post_exam_teacher, normal, post_exam])
        self.db.flush()
        self.db.add_all([
            TimetableLesson(
                version_id=normal.id, weekday=4, period=1, class_code="1A",
                subject="Normal lesson", teachers_json=json.dumps([normal_teacher.id]),
            ),
            TimetableLesson(
                version_id=post_exam.id, weekday=0, period=2, class_code="1A",
                subject="Post-exam lesson", teachers_json=json.dumps([post_exam_teacher.id]),
            ),
        ])
        absence = AbsenceCase(
            professor_id=normal_teacher.id, data=date(2026, 6, 19),
            periods_json="[1]", status="open",
        )
        self.db.add(absence)
        self.db.commit()

        task = analyze_absence(self.db, absence)["tasks"][0]

        self.assertEqual(task["status"], "unresolved")
        self.assertIsNone(task["recommended"])
        self.assertEqual(task["alternatives"], [])

    def test_existing_timetable_special_subjects_can_be_updated(self):
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), effective_to=date(2026, 8, 31),
            class_filename="classes.xls", teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add(version)
        self.db.flush()
        lessons = [
            TimetableLesson(
                version_id=version.id, weekday=0, period=index, class_code="1A",
                subject=subject, teachers_json="[]",
            )
            for index, subject in enumerate(("體育", "中文"), 1)
        ]
        self.db.add_all(lessons)
        self.db.commit()

        update_timetable_version(
            version.id,
            UpdateTimetableRequest(
                effective_from=version.effective_from, effective_to=version.effective_to,
                special_subjects=["體育"],
            ),
            self.db,
            SimpleNamespace(username="admin"),
        )

        self.assertTrue(lessons[0].special)
        self.assertFalse(lessons[1].special)
        row = list_timetable_versions(self.db)[0]
        self.assertEqual(row["subjects"], ["中文", "體育"])
        self.assertEqual(row["special_subjects"], ["體育"])

    @patch("routes.rescheduling.hong_kong_now", return_value=datetime(2026, 8, 10, 12, 0))
    def test_special_subject_change_marks_only_unfinished_cross_day_adjustments(self, _now):
        version = TimetableVersion(
            effective_from=date(2026, 8, 1), effective_to=date(2026, 8, 31),
            class_filename="classes.xls", teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add(version)
        self.db.flush()
        lesson = TimetableLesson(
            version_id=version.id, weekday=0, period=1, class_code="1A",
            subject="體育", teachers_json="[]",
        )
        future = ScheduleAdjustment(kind="direct_swap", status="confirmed", locked=True)
        past = ScheduleAdjustment(kind="direct_swap", status="confirmed", locked=True)
        self.db.add_all([lesson, future, past])
        self.db.flush()
        self.db.add_all([
            ScheduleAdjustmentLeg(
                adjustment_id=future.id, lesson_id=lesson.id, class_code="1A", subject="體育",
                teachers_json="[]", from_date=date(2026, 8, 9), from_period=1,
                to_date=date(2026, 8, 10), to_period=8,
            ),
            ScheduleAdjustmentLeg(
                adjustment_id=past.id, lesson_id=lesson.id, class_code="1A", subject="體育",
                teachers_json="[]", from_date=date(2026, 8, 9), from_period=1,
                to_date=date(2026, 8, 10), to_period=1,
            ),
        ])
        self.db.commit()

        result = update_timetable_version(
            version.id,
            UpdateTimetableRequest(
                effective_from=version.effective_from, effective_to=version.effective_to,
                special_subjects=["體育"],
            ),
            self.db,
            SimpleNamespace(username="admin"),
        )

        self.db.refresh(future)
        self.db.refresh(past)
        self.assertEqual(result["review_required_count"], 1)
        self.assertTrue(future.needs_review)
        self.assertFalse(past.needs_review)

    def test_import_preview_can_be_discarded(self):
        preview = TimetableImportPreview(
            id="discard-me", payload="{}", class_filename="classes.xls",
            teacher_filename="teachers.xlsx", created_by="admin",
        )
        self.db.add(preview)
        self.db.commit()

        discard_import_preview("discard-me", self.db, SimpleNamespace(username="admin"))

        self.assertIsNone(self.db.get(TimetableImportPreview, "discard-me"))

    def test_test_records_can_be_updated_and_purged_transactionally(self):
        teacher = Professor(nom="A老師", actiu=True)
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), effective_to=date(2026, 8, 31),
            class_filename="classes.xls", teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([teacher, version])
        self.db.flush()
        lesson = TimetableLesson(
            version_id=version.id, weekday=0, period=1, class_code="1A",
            subject="中文", teachers_json=json.dumps([teacher.id]),
        )
        absence = AbsenceCase(
            professor_id=teacher.id, data=date(2026, 8, 11), periods_json="[1]", status="resolved",
        )
        adjustment = ScheduleAdjustment(
            absence_case_id=None, kind="direct_swap", status="confirmed", locked=True,
        )
        self.db.add_all([lesson, absence, adjustment])
        self.db.flush()
        adjustment.absence_case_id = absence.id
        self.db.add(ScheduleAdjustmentLeg(
            adjustment_id=adjustment.id, lesson_id=lesson.id, class_code="1A", subject="中文",
            teachers_json=json.dumps([teacher.id]), from_date=date(2026, 8, 11), from_period=1,
            to_date=date(2026, 8, 11), to_period=2,
        ))
        self.db.commit()
        user = SimpleNamespace(username="admin")

        reason_updated = update_absence(
            absence.id,
            AbsenceCreateRequest(
                professor_id=teacher.id, data=date(2026, 8, 11), periods=[1],
                reason_type="other", reason_detail="更正原因",
            ),
            self.db, user,
        )
        self.assertIsNotNone(self.db.get(ScheduleAdjustment, adjustment.id))
        self.assertEqual(absence.status, "resolved")
        self.assertEqual((reason_updated["reason_type"], reason_updated["reason_detail"]), ("other", "更正原因"))

        updated = update_absence(
            absence.id,
            AbsenceCreateRequest(professor_id=teacher.id, data=date(2026, 8, 11), periods=[1, 2], reason_type="personal"),
            self.db, user,
        )
        self.assertIsNotNone(self.db.get(ScheduleAdjustment, adjustment.id))
        self.assertEqual(json.loads(absence.periods_json), [1, 2])
        self.assertEqual(updated["reason_type"], "personal")
        record = list_records(scope="all", page=1, page_size=20, db=self.db)["items"][0]
        self.assertEqual((record["reason_type"], record["reason_detail"]), ("personal", None))

        with self.assertRaises(HTTPException) as identity_change:
            update_absence(
                absence.id,
                AbsenceCreateRequest(
                    professor_id=teacher.id, data=date(2026, 8, 12), periods=[2, 3], reason_type="personal"
                ),
                self.db, user,
            )
        self.assertEqual(identity_change.exception.status_code, 409)
        update_adjustment(adjustment.id, UpdateAdjustmentRequest(reason="測試原因"), self.db, user)
        self.assertEqual(adjustment.reason, "測試原因")
        with self.assertRaises(HTTPException) as confirmed_delete:
            delete_adjustment(adjustment.id, self.db, user)
        self.assertEqual(confirmed_delete.exception.status_code, 409)
        purge_absence(absence.id, self.db, user)
        self.assertIsNone(self.db.get(AbsenceCase, absence.id))

    def test_editing_periods_recomputes_status_with_confirmed_cover(self):
        absent = Professor(nom="缺席老師", actiu=True)
        replacement = Professor(nom="代課老師", actiu=True)
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), effective_to=date(2026, 8, 31),
            class_filename="classes.xls", teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([absent, replacement, version])
        self.db.flush()
        lessons = [
            TimetableLesson(
                version_id=version.id, weekday=0, period=period, class_code="1A",
                subject=subject, teachers_json=json.dumps([absent.id]),
            )
            for period, subject in ((1, "數學"), (2, "中文"))
        ]
        absence = AbsenceCase(
            professor_id=absent.id, data=date(2026, 8, 10), periods_json="[1, 2]",
            reason_type="sick", status="open", created_by="admin",
        )
        self.db.add_all([*lessons, absence])
        self.db.flush()
        adjustment = ScheduleAdjustment(
            absence_case_id=absence.id, kind="emergency_cover", status="confirmed", locked=True,
        )
        self.db.add(adjustment)
        self.db.flush()
        self.db.add(ScheduleAdjustmentLeg(
            adjustment_id=adjustment.id, lesson_id=lessons[0].id,
            class_code="1A", subject="數學", teachers_json=json.dumps([absent.id]),
            from_date=absence.data, from_period=1, to_date=absence.data, to_period=1,
            replaced_teacher_id=absent.id, replacement_teacher_id=replacement.id,
        ))
        self.db.commit()

        update_absence(
            absence.id,
            AbsenceCreateRequest(
                professor_id=absent.id, data=absence.data, periods=[1], reason_type="sick",
            ),
            self.db,
            SimpleNamespace(username="admin"),
        )

        self.assertEqual(absence.status, "resolved")
        self.assertEqual(self.db.get(ScheduleAdjustment, adjustment.id).status, "confirmed")
        self.assertEqual(analyze_absence(self.db, absence)["tasks"], [])

        update_absence(
            absence.id,
            AbsenceCreateRequest(
                professor_id=absent.id, data=absence.data, periods=[1, 2], reason_type="sick",
            ),
            self.db,
            SimpleNamespace(username="admin"),
        )
        self.assertEqual(absence.status, "open")

    @patch("rescheduling_service.parse_class_workbook")
    @patch("rescheduling_service.parse_teacher_workbook")
    @patch("rescheduling_service.teacher_names_from_workbook")
    def test_import_preview_reports_exact_mismatch(self, names_parser, teacher_parser, class_parser):
        names_parser.return_value = ["A老師"]
        teacher_parser.return_value = ([{
            "teacher": "A老師", "weekday": 0, "period": 1,
            "class_code": "1B", "subject": "中文",
        }], ["A老師"])
        class_parser.return_value = ([{
            "weekday": 0, "period": 1, "class_code": "1A",
            "subject": "中文", "teachers": ["A老師"],
        }], [(505, 540)])

        result = build_import_preview(b"class", b"teacher")

        self.assertEqual(result["summary"]["blocked_lessons"], 1)
        self.assertEqual(result["issues"][0]["code"], "teacher_slot_mismatch")
        self.assertEqual(result["issues"][0]["teacher_workbook"], "1B 中文")
        resolution_id = result["issues"][0]["resolution_id"]

        class_result = apply_import_resolutions(deepcopy(result), {resolution_id: "class"})
        self.assertTrue(class_result["lessons"][0]["movable"])
        self.assertEqual(class_result["teacher_slots"][0]["class_code"], "1A")

        teacher_result = apply_import_resolutions(deepcopy(result), {resolution_id: "teacher"})
        self.assertEqual(teacher_result["lessons"][0]["teachers"], [])
        self.assertFalse(teacher_result["lessons"][0]["movable"])
        self.assertEqual(teacher_result["teacher_slots"][0]["class_code"], "1B")

    @patch("rescheduling_service.parse_class_workbook")
    @patch("rescheduling_service.parse_teacher_workbook")
    @patch("rescheduling_service.teacher_names_from_workbook")
    def test_import_preview_can_add_teacher_workbook_extra_as_co_teacher(
        self, names_parser, teacher_parser, class_parser,
    ):
        names_parser.return_value = ["A老師", "B老師"]
        teacher_parser.return_value = ([
            {"teacher": "A老師", "weekday": 3, "period": 2, "class_code": "3C", "subject": "中文"},
            {"teacher": "B老師", "weekday": 3, "period": 2, "class_code": "3C", "subject": "中文"},
        ], ["A老師", "B老師"])
        class_parser.return_value = ([{
            "weekday": 3, "period": 2, "class_code": "3C",
            "subject": "中文", "teachers": ["B老師"],
        }], [(540, 575)])

        result = build_import_preview(b"class", b"teacher")

        self.assertEqual(len(result["issues"]), 1)
        self.assertEqual(result["issues"][0]["code"], "teacher_extra_assignment")
        resolution_id = result["issues"][0]["resolution_id"]

        teacher_result = apply_import_resolutions(deepcopy(result), {resolution_id: "teacher"})
        self.assertEqual(teacher_result["lessons"][0]["teachers"], ["B老師", "A老師"])
        self.assertTrue(teacher_result["lessons"][0]["movable"])

        class_result = apply_import_resolutions(deepcopy(result), {resolution_id: "class"})
        self.assertEqual(class_result["lessons"][0]["teachers"], ["B老師"])
        self.assertEqual([slot["teacher"] for slot in class_result["teacher_slots"]], ["B老師"])

    def test_import_review_decisions_can_be_saved_partially(self):
        preview = TimetableImportPreview(
            id="preview-1",
            class_filename="classes.xls",
            teacher_filename="teachers.xlsx",
            payload=json.dumps({
                "issues": [
                    {"severity": "review", "resolution_id": "first"},
                    {"severity": "review", "resolution_id": "second"},
                ]
            }, ensure_ascii=False),
        )
        self.db.add(preview)
        self.db.commit()

        result = save_import_resolutions(
            preview.id,
            SaveImportResolutionsRequest(resolutions={"first": "teacher"}),
            self.db,
            SimpleNamespace(username="admin"),
        )

        self.assertEqual(result["saved_resolutions"], {"first": "teacher"})
        self.assertEqual((result["confirmed_count"], result["remaining_count"]), (1, 1))
        saved_payload = json.loads(self.db.get(TimetableImportPreview, preview.id).payload)
        self.assertEqual(saved_payload["saved_resolutions"], {"first": "teacher"})

    def _confirmed_swap_fixture(self):
        teachers = [Professor(nom=f"{name}老師", actiu=True) for name in "ABCD"]
        self.db.add_all(teachers)
        self.db.flush()
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), class_filename="classes.xls",
            teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add(version)
        self.db.flush()
        lessons = [
            TimetableLesson(
                version_id=version.id, weekday=0, period=period, class_code="1A",
                subject=subject, teachers_json=json.dumps([teacher.id]),
            )
            for teacher, period, subject in zip(teachers, (1, 3, 5, 2), ("中文", "英文", "數學", "常識"))
        ]
        absence = AbsenceCase(
            professor_id=teachers[0].id, data=date(2026, 8, 10),
            periods_json="[1, 3]", reason_type="sick", status="open",
        )
        self.db.add_all([*lessons, absence])
        self.db.flush()
        adjustment = ScheduleAdjustment(
            absence_case_id=absence.id, kind="direct_swap", status="confirmed", locked=True,
        )
        self.db.add(adjustment)
        self.db.flush()
        self.db.add_all([
            ScheduleAdjustmentLeg(
                adjustment_id=adjustment.id, lesson_id=lessons[0].id, class_code="1A", subject="中文",
                teachers_json=json.dumps([teachers[0].id]), from_date=date(2026, 8, 10), from_period=1,
                to_date=date(2026, 8, 10), to_period=3,
            ),
            ScheduleAdjustmentLeg(
                adjustment_id=adjustment.id, lesson_id=lessons[1].id, class_code="1A", subject="英文",
                teachers_json=json.dumps([teachers[1].id]), from_date=date(2026, 8, 10), from_period=3,
                to_date=date(2026, 8, 10), to_period=1,
            ),
        ])
        self.db.commit()
        return teachers, lessons, absence, adjustment

    def test_confirmed_swap_changes_effective_timetable_without_changing_base(self):
        teacher_a = Professor(nom="A老師", actiu=True)
        teacher_b = Professor(nom="B老師", actiu=True)
        self.db.add_all([teacher_a, teacher_b])
        self.db.flush()
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), class_filename="classes.xls",
            teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add(version)
        self.db.flush()
        lesson_a = TimetableLesson(
            version_id=version.id, weekday=0, period=1, class_code="1A",
            subject="中文", teachers_json=json.dumps([teacher_a.id]),
        )
        lesson_b = TimetableLesson(
            version_id=version.id, weekday=0, period=4, class_code="1A",
            subject="英文", teachers_json=json.dumps([teacher_b.id]),
        )
        self.db.add_all([lesson_a, lesson_b])
        self.db.flush()
        adjustment = ScheduleAdjustment(kind="direct_swap", status="confirmed", locked=True)
        self.db.add(adjustment)
        self.db.flush()
        self.db.add_all([
            ScheduleAdjustmentLeg(
                adjustment_id=adjustment.id, lesson_id=lesson_a.id, class_code="1A",
                subject="中文", teachers_json=json.dumps([teacher_a.id]),
                from_date=date(2026, 8, 10), from_period=1,
                to_date=date(2026, 8, 10), to_period=4,
            ),
            ScheduleAdjustmentLeg(
                adjustment_id=adjustment.id, lesson_id=lesson_b.id, class_code="1A",
                subject="英文", teachers_json=json.dumps([teacher_b.id]),
                from_date=date(2026, 8, 10), from_period=4,
                to_date=date(2026, 8, 10), to_period=1,
            ),
        ])
        self.db.commit()

        effective = effective_occurrences(self.db, date(2026, 8, 10), date(2026, 8, 10))
        by_period = {row["period"]: row for row in effective}
        self.assertEqual(by_period[1]["teachers"], [teacher_b.id])
        self.assertEqual(by_period[4]["teachers"], [teacher_a.id])
        self.assertEqual(self.db.get(TimetableLesson, lesson_a.id).period, 1)
        self.assertTrue(by_period[1]["locked"])
        payload = get_effective_timetable(data=date(2026, 8, 10), db=self.db)
        movements = {(row["subject"], row["from_period"], row["to_period"]) for row in payload["lessons"]}
        self.assertEqual(movements, {("中文", 1, 4), ("英文", 4, 1)})

    def test_confirmed_moved_lesson_is_repaired_locally_and_keeps_history(self):
        teachers, lessons, absence, first_adjustment = self._confirmed_swap_fixture()
        now = datetime(2026, 8, 10, 9, 10)
        period_starts = {item["period"]: item["start"] for item in get_period_times(self.db)}

        analysis = analyze_absence(self.db, absence, now=now, period_starts=period_starts)
        task = analysis["tasks"][0]

        self.assertEqual((len(analysis["tasks"]), task["target"]["period"]), (1, 3))
        self.assertEqual(task["recommended"]["kind"], "direct_swap")
        candidate_lesson_ids = {leg["lesson_id"] for leg in task["recommended"]["legs"]}
        self.assertEqual(candidate_lesson_ids, {lessons[0].id, lessons[2].id})
        self.assertTrue(all(
            not (leg["from_date"] == "2026-08-10" and leg["from_period"] == 2)
            for candidate in task["alternatives"] for leg in candidate["legs"]
        ))

        with patch("routes.rescheduling.hong_kong_now", return_value=now):
            saved = confirm_adjustment(
                ConfirmRequest(
                    absence_case_id=absence.id,
                    candidate_id=task["recommended"]["id"],
                    expected_revision=analysis["revision"],
                ),
                self.db,
                SimpleNamespace(username="admin"),
            )

        effective = effective_occurrences(self.db, absence.data, absence.data)
        by_period = {row["period"]: row["subject"] for row in effective if row["class_code"] == "1A"}
        self.assertEqual(by_period, {1: "英文", 2: "常識", 3: "數學", 5: "中文"})
        self.assertEqual(len(effective), 4)
        self.assertEqual(self.db.query(ScheduleAdjustment).filter_by(status="confirmed").count(), 2)
        self.db.refresh(absence)
        self.assertEqual(absence.status, "resolved")

        with patch("routes.rescheduling.hong_kong_now", return_value=datetime(2026, 8, 10, 8, 0)):
            with self.assertRaises(HTTPException) as upstream:
                revert_adjustment(first_adjustment.id, self.db, SimpleNamespace(username="admin"))
        self.assertIn("後續安排", upstream.exception.detail)

        with patch("routes.rescheduling.hong_kong_now", return_value=now):
            revert_adjustment(saved["id"], self.db, SimpleNamespace(username="admin"))
            with self.assertRaises(HTTPException) as started:
                revert_adjustment(first_adjustment.id, self.db, SimpleNamespace(username="admin"))
        self.assertIn("已開始或已過去", started.exception.detail)

    def test_batch_append_keeps_confirmed_adjustment_and_analyzes_new_period(self):
        _teachers, _lessons, absence, adjustment = self._confirmed_swap_fixture()
        absence.periods_json = "[1]"
        absence.status = "resolved"
        absence.created_by = "admin"
        self.db.commit()

        result = create_absences_batch(
            AbsenceBatchCreateRequest(items=[AbsenceCreateRequest(
                professor_id=absence.professor_id,
                data=absence.data,
                periods=[3],
                reason_type="personal",
            )]),
            self.db,
            SimpleNamespace(username="admin"),
        )

        self.assertEqual(json.loads(absence.periods_json), [1, 3])
        self.assertEqual(absence.reason_type, "sick")
        self.assertEqual(absence.status, "open")
        self.assertEqual(self.db.get(ScheduleAdjustment, adjustment.id).status, "confirmed")
        self.assertEqual([task["target"]["period"] for task in result["tasks"]], [3])

    def test_started_effective_lesson_is_manual_only(self):
        _, _, absence, _ = self._confirmed_swap_fixture()
        period_starts = {item["period"]: item["start"] for item in get_period_times(self.db)}

        task = analyze_absence(
            self.db, absence, now=datetime(2026, 8, 10, 10, 0), period_starts=period_starts
        )["tasks"][0]

        self.assertEqual(task["status"], "unresolved")
        self.assertEqual(task["blocking_reason"], "started")
        self.assertEqual(task["alternatives"], [])

    def test_many_confirmed_moves_overlay_without_duplicates(self):
        teachers = [Professor(nom=f"{name}老師", actiu=True) for name in "AB"]
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), class_filename="classes.xls",
            teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([*teachers, version])
        self.db.flush()
        lessons = [
            TimetableLesson(
                version_id=version.id, weekday=0, period=period, class_code="1A",
                subject=subject, teachers_json=json.dumps([teacher.id]),
            )
            for teacher, period, subject in zip(teachers, (1, 2), ("中文", "英文"))
        ]
        self.db.add_all(lessons)
        self.db.flush()
        positions = [1, 2]
        for _ in range(51):
            adjustment = ScheduleAdjustment(kind="direct_swap", status="confirmed", locked=True)
            self.db.add(adjustment)
            self.db.flush()
            self.db.add_all([
                ScheduleAdjustmentLeg(
                    adjustment_id=adjustment.id, lesson_id=lesson.id, class_code="1A", subject=lesson.subject,
                    teachers_json=lesson.teachers_json, from_date=date(2026, 8, 10), from_period=positions[index],
                    to_date=date(2026, 8, 10), to_period=positions[1 - index],
                )
                for index, lesson in enumerate(lessons)
            ])
            positions.reverse()
        self.db.commit()

        effective = effective_occurrences(self.db, date(2026, 8, 10), date(2026, 8, 10))

        self.assertEqual(len(effective), 2)
        self.assertEqual({row["period"]: row["subject"] for row in effective}, {1: "英文", 2: "中文"})
        self.assertEqual(self.db.query(ScheduleAdjustmentLeg).count(), 102)

    def test_effective_timetable_uses_version_for_each_date(self):
        teacher = Professor(nom="A老師", actiu=True)
        self.db.add(teacher)
        self.db.flush()
        old_version = TimetableVersion(
            effective_from=date(2026, 8, 10), class_filename="old.xls",
            teacher_filename="old.xlsx", active=False,
        )
        new_version = TimetableVersion(
            effective_from=date(2026, 8, 17), class_filename="new.xls",
            teacher_filename="new.xlsx", active=True,
        )
        self.db.add_all([old_version, new_version])
        self.db.flush()
        self.db.add_all([
            TimetableLesson(
                version_id=old_version.id, weekday=0, period=1, class_code="1A",
                subject="舊課表", teachers_json=json.dumps([teacher.id]),
            ),
            TimetableLesson(
                version_id=new_version.id, weekday=0, period=1, class_code="1A",
                subject="新課表", teachers_json=json.dumps([teacher.id]),
            ),
        ])
        self.db.commit()

        rows = effective_occurrences(self.db, date(2026, 8, 10), date(2026, 8, 17))
        subjects = {(row["date"], row["subject"]) for row in rows}
        self.assertIn((date(2026, 8, 10), "舊課表"), subjects)
        self.assertIn((date(2026, 8, 17), "新課表"), subjects)
        self.assertNotIn((date(2026, 8, 10), "新課表"), subjects)

    def test_global_choice_does_not_reuse_same_teacher_slot(self):
        shared_leg = {
            "occurrence_id": "lesson-a", "lesson_id": 1, "class_code": "1A",
            "subject": "中文", "teachers": [1], "from_date": "2026-08-10",
            "from_period": 1, "to_date": "2026-08-10", "to_period": 4,
        }
        first = {"id": "first", "day_distance": 0, "moved_lessons": 2,
                 "legs": [shared_leg]}
        second = {"id": "second", "day_distance": 0, "moved_lessons": 2,
                  "legs": [{**shared_leg, "occurrence_id": "lesson-b"}]}
        chosen = choose_global([[first], [second]])
        self.assertEqual(len(chosen), 1)

    def test_analysis_prefers_same_day_direct_swap(self):
        teacher_a = Professor(nom="A老師", actiu=True)
        teacher_b = Professor(nom="B老師", actiu=True)
        self.db.add_all([teacher_a, teacher_b])
        self.db.flush()
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), class_filename="classes.xls",
            teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add(version)
        self.db.flush()
        self.db.add_all([
            TimetableLesson(
                version_id=version.id, weekday=0, period=1, class_code="1A",
                subject="中文", teachers_json=json.dumps([teacher_a.id]),
            ),
            TimetableLesson(
                version_id=version.id, weekday=0, period=4, class_code="1A",
                subject="英文", teachers_json=json.dumps([teacher_b.id]),
            ),
        ])
        absence = AbsenceCase(
            professor_id=teacher_a.id, data=date(2026, 8, 10),
            periods_json="[1]", status="open",
        )
        self.db.add(absence)
        self.db.commit()

        analysis = analyze_absence(self.db, absence)
        self.assertEqual(analysis["resolved_count"], 1)
        self.assertEqual(analysis["tasks"][0]["recommended"]["kind"], "direct_swap")
        self.assertEqual(analysis["tasks"][0]["recommended"]["day_distance"], 0)

    def test_analysis_deprioritizes_swap_that_breaks_core_consecutive_lessons(self):
        teacher_a = Professor(nom="A老師", actiu=True)
        teacher_b = Professor(nom="B老師", actiu=True)
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), class_filename="classes.xls",
            teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([teacher_a, teacher_b, version])
        self.db.flush()
        self.db.add_all([
            TimetableLesson(
                version_id=version.id, weekday=0, period=1, class_code="1A",
                subject="常識", teachers_json=json.dumps([teacher_a.id]),
            ),
        ] + [
            TimetableLesson(
                version_id=version.id, weekday=0, period=period, class_code="1A",
                subject="中文", teachers_json=json.dumps([teacher_b.id]),
            ) for period in (2, 3)
        ])
        absence = AbsenceCase(
            professor_id=teacher_a.id, data=date(2026, 8, 10), periods_json="[1]", status="open",
        )
        self.db.add(absence)
        self.db.commit()

        task = analyze_absence(self.db, absence)["tasks"][0]

        self.assertEqual(task["recommended"]["breaks_consecutive_lessons"], 0)
        self.assertTrue(any(candidate["breaks_consecutive_lessons"] for candidate in task["alternatives"]))

    def test_analysis_deprioritizes_cross_day_special_lessons(self):
        teachers = [Professor(nom=f"{name}老師", actiu=True) for name in "ABC"]
        self.db.add_all(teachers)
        self.db.flush()
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), class_filename="classes.xls",
            teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add(version)
        self.db.flush()
        self.db.add_all([
            TimetableLesson(
                version_id=version.id, weekday=0, period=1, class_code="1A",
                subject="中文", teachers_json=json.dumps([teachers[0].id]),
            ),
            TimetableLesson(
                version_id=version.id, weekday=1, period=2, class_code="1A",
                subject="體育", teachers_json=json.dumps([teachers[1].id]), special=True,
            ),
            TimetableLesson(
                version_id=version.id, weekday=2, period=2, class_code="1A",
                subject="數學", teachers_json=json.dumps([teachers[2].id]),
            ),
        ])
        absence = AbsenceCase(
            professor_id=teachers[0].id, data=date(2026, 8, 10),
            periods_json="[1]", status="open",
        )
        self.db.add(absence)
        self.db.commit()

        candidate = analyze_absence(self.db, absence)["tasks"][0]["recommended"]

        self.assertEqual(candidate["completion_date"], "2026-08-12")
        self.assertEqual(candidate["special_cross_day_moves"], 0)

    def test_analysis_uses_three_lesson_cycle_when_direct_swaps_conflict(self):
        teachers = [Professor(nom=f"{name}老師", actiu=True) for name in "ABC"]
        self.db.add_all(teachers)
        self.db.flush()
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), class_filename="classes.xls",
            teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add(version)
        self.db.flush()
        rows = [
            (1, "1A", "中文", teachers[0]), (2, "1A", "英文", teachers[1]),
            (3, "1A", "數學", teachers[2]),
            # Block A<->B because B is busy at period 1; block A<->C because A is busy at period 3.
            (1, "2A", "英文", teachers[1]), (3, "2B", "中文", teachers[0]),
        ]
        for period, class_code, subject, teacher in rows:
            self.db.add(TimetableLesson(
                version_id=version.id, weekday=0, period=period, class_code=class_code,
                subject=subject, teachers_json=json.dumps([teacher.id]),
            ))
        absence = AbsenceCase(
            professor_id=teachers[0].id, data=date(2026, 8, 10), periods_json="[1]", status="open",
        )
        self.db.add(absence)
        self.db.commit()

        analysis = analyze_absence(self.db, absence)
        self.assertEqual(analysis["tasks"][0]["recommended"]["kind"], "three_cycle")
        self.assertEqual(len(analysis["tasks"][0]["recommended"]["legs"]), 3)

    def test_configured_cycle_limit_enables_four_lesson_recommendation(self):
        teachers = [Professor(nom=f"{name}老師", actiu=True) for name in "ABCD"]
        self.db.add_all(teachers)
        self.db.flush()
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), class_filename="classes.xls",
            teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add(version)
        self.db.flush()
        main_lessons = [
            (1, "中文", teachers[0]), (2, "英文", teachers[1]),
            (3, "數學", teachers[2]), (4, "科學", teachers[3]),
        ]
        blockers = [
            (1, "2A", teachers[1]), (1, "2B", teachers[2]),
            (3, "2C", teachers[0]), (4, "2D", teachers[0]), (4, "2E", teachers[1]),
        ]
        self.db.add_all([
            TimetableLesson(
                version_id=version.id, weekday=0, period=period, class_code="1A",
                subject=subject, teachers_json=json.dumps([teacher.id]),
            ) for period, subject, teacher in main_lessons
        ] + [
            TimetableLesson(
                version_id=version.id, weekday=0, period=period, class_code=class_code,
                subject="阻擋", teachers_json=json.dumps([teacher.id]),
            ) for period, class_code, teacher in blockers
        ])
        absence = AbsenceCase(
            professor_id=teachers[0].id, data=date(2026, 8, 10), periods_json="[1]", status="open",
        )
        self.db.add(absence)
        self.db.commit()

        self.assertEqual(analyze_absence(self.db, absence)["tasks"][0]["status"], "unresolved")
        saved = update_rescheduling_config(
            ReschedulingConfigRequest(max_cycle_lessons=4),
            self.db,
            SimpleNamespace(username="admin"),
        )
        candidate = analyze_absence(self.db, absence)["tasks"][0]["recommended"]

        self.assertEqual(saved["max_cycle_lessons"], 4)
        self.assertEqual(candidate["kind"], "three_cycle")
        self.assertEqual(len(candidate["legs"]), 4)

    def test_emergency_cover_is_same_subject_and_only_when_no_swap_exists(self):
        absent_teacher = Professor(nom="A老師", actiu=True)
        same_subject_teacher = Professor(nom="C老師", actiu=True)
        self.db.add_all([absent_teacher, same_subject_teacher])
        self.db.flush()
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), class_filename="classes.xls",
            teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add(version)
        self.db.flush()
        self.db.add_all([
            TimetableLesson(
                version_id=version.id, weekday=0, period=1, class_code="1A",
                subject="中文", teachers_json=json.dumps([absent_teacher.id]),
            ),
            TimetableLesson(
                version_id=version.id, weekday=0, period=2, class_code="2A",
                subject="中文", teachers_json=json.dumps([same_subject_teacher.id]),
            ),
        ])
        absence = AbsenceCase(
            professor_id=absent_teacher.id, data=date(2026, 8, 10), periods_json="[1]", status="open",
        )
        self.db.add(absence)
        self.db.commit()

        analysis = analyze_absence(self.db, absence)
        candidate = analysis["tasks"][0]["recommended"]
        self.assertEqual(candidate["kind"], "emergency_cover")
        self.assertEqual(candidate["legs"][0]["replacement_teacher_id"], same_subject_teacher.id)

    def test_emergency_cover_deprioritizes_new_consecutive_teaching(self):
        absent = Professor(nom="A缺席", actiu=True)
        adjacent = Professor(nom="B相鄰有課", actiu=True)
        clear = Professor(nom="C前後空堂", actiu=True)
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), class_filename="classes.xls",
            teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([absent, adjacent, clear, version])
        self.db.flush()
        self.db.add_all([
            TimetableLesson(
                version_id=version.id, weekday=0, period=1, class_code="1A",
                subject="中文", teachers_json=json.dumps([absent.id]),
            ),
            TimetableLesson(
                version_id=version.id, weekday=0, period=2, class_code="2A",
                subject="中文", teachers_json=json.dumps([adjacent.id]),
            ),
            TimetableLesson(
                version_id=version.id, weekday=1, period=1, class_code="3A",
                subject="中文", teachers_json=json.dumps([clear.id]),
            ),
        ])
        absence = AbsenceCase(
            professor_id=absent.id, data=date(2026, 8, 10), periods_json="[1]", status="open",
        )
        self.db.add(absence)
        self.db.commit()

        task = analyze_absence(self.db, absence)["tasks"][0]

        self.assertEqual(task["recommended"]["legs"][0]["replacement_teacher_id"], clear.id)
        self.assertEqual(task["recommended"]["new_consecutive_classes"], 0)
        self.assertTrue(any(candidate["new_consecutive_classes"] for candidate in task["alternatives"]))

    def test_common_planning_slot_blocks_emergency_cover(self):
        absent_teacher = Professor(nom="A老師", actiu=True)
        planning_teacher = Professor(nom="C老師", actiu=True)
        self.db.add_all([absent_teacher, planning_teacher])
        self.db.flush()
        version = TimetableVersion(
            effective_from=date(2026, 8, 10), class_filename="classes.xls",
            teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add(version)
        self.db.flush()
        self.db.add_all([
            TimetableLesson(
                version_id=version.id, weekday=0, period=1, class_code="1A",
                subject="中文", teachers_json=json.dumps([absent_teacher.id]),
            ),
            TimetableLesson(
                version_id=version.id, weekday=0, period=2, class_code="2A",
                subject="中文", teachers_json=json.dumps([planning_teacher.id]),
            ),
            TimetableTeacherSlot(
                version_id=version.id, professor_id=planning_teacher.id, weekday=0,
                period=1, class_code="", subject="共同備課",
            ),
        ])
        absence = AbsenceCase(
            professor_id=absent_teacher.id, data=date(2026, 8, 10), periods_json="[1]", status="open",
        )
        self.db.add(absence)
        self.db.commit()

        analysis = analyze_absence(self.db, absence)

        self.assertEqual(analysis["tasks"][0]["status"], "unresolved")
        self.assertEqual(analysis["tasks"][0]["alternatives"], [])

    def test_long_term_leave_blocks_every_period_in_date_range(self):
        teacher = Professor(nom="A老師", actiu=True)
        self.db.add(teacher)
        self.db.flush()
        self.db.add(ProfessorBaixa(
            professor=teacher.nom,
            data_inici=date(2026, 8, 12),
            data_final=date(2026, 8, 14),
            motiu="maternity",
        ))
        self.db.commit()

        keys = absence_keys(self.db, date(2026, 8, 11), date(2026, 8, 15))

        self.assertNotIn((teacher.id, date(2026, 8, 11), 1), keys)
        self.assertIn((teacher.id, date(2026, 8, 12), 1), keys)
        self.assertIn((teacher.id, date(2026, 8, 14), 9), keys)
        self.assertNotIn((teacher.id, date(2026, 8, 15), 9), keys)

    @patch("routes.rescheduling.hong_kong_today", return_value=date(2026, 8, 11))
    def test_records_are_split_by_hong_kong_today_and_paginated(self, _today):
        teacher = Professor(nom="A老師", actiu=True)
        self.db.add(teacher)
        self.db.flush()
        cases = []
        for day in (date(2026, 8, 10), date(2026, 8, 11), date(2026, 8, 12), date(2026, 8, 13)):
            cases.append(AbsenceCase(
                professor_id=teacher.id, data=day, periods_json="[1]", status="open",
            ))
        self.db.add_all(cases)
        self.db.flush()
        self.db.add(AbsenceCase(
            professor_id=teacher.id, data=date(2026, 8, 12), periods_json="[2]", status="cancelled",
        ))
        self.db.add(ScheduleAdjustment(
            absence_case_id=cases[2].id, kind="direct_swap", status="reverted", locked=False,
        ))
        cases[3].status = "resolved"
        self.db.add(ScheduleAdjustment(
            absence_case_id=cases[3].id, kind="emergency_cover", status="confirmed", locked=True,
        ))
        self.db.commit()

        today = list_records(scope="today", page=1, page_size=20, db=self.db)
        future = list_records(scope="future", page=1, page_size=1, db=self.db)
        past = list_records(scope="past", page=1, page_size=20, db=self.db)
        filtered = list_records(
            scope="all", page=1, page_size=20, date_from=date(2026, 8, 11),
            date_to=date(2026, 8, 12), q="A老師", status="open", db=self.db,
        )
        covers = list_records(scope="all", page=1, page_size=20, status="completed", kind="cover", db=self.db)

        self.assertEqual([item["date"] for item in today["items"]], ["2026-08-11"])
        self.assertEqual(future["total"], 2)
        self.assertEqual(future["pages"], 2)
        self.assertEqual(future["items"][0]["date"], "2026-08-12")
        self.assertEqual(future["items"][0]["adjustments"], [])
        self.assertEqual([item["date"] for item in past["items"]], ["2026-08-10"])
        self.assertEqual([item["date"] for item in filtered["items"]], ["2026-08-12", "2026-08-11"])
        self.assertEqual([item["date"] for item in covers["items"]], ["2026-08-13"])

    def test_emergency_cover_has_single_formatted_export_remark(self):
        absent = Professor(nom="劉慧妍", actiu=True)
        replacement = Professor(nom="吳淑君", actiu=True)
        version = TimetableVersion(
            effective_from=date(2026, 8, 1), effective_to=date(2027, 7, 31),
            class_filename="classes.xls", teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([absent, replacement, version])
        self.db.flush()
        lesson = TimetableLesson(
            version_id=version.id, weekday=0, period=1, class_code="2B", subject="數學",
            teachers_json=json.dumps([absent.id]),
        )
        absence = AbsenceCase(
            professor_id=absent.id, data=date(2026, 8, 10), periods_json="[1]",
            reason_type="sick", status="resolved",
        )
        self.db.add_all([lesson, absence])
        self.db.flush()
        adjustment = ScheduleAdjustment(
            absence_case_id=absence.id, kind="emergency_cover", status="confirmed", locked=True,
        )
        self.db.add(adjustment)
        self.db.flush()
        self.db.add(ScheduleAdjustmentLeg(
            adjustment_id=adjustment.id, lesson_id=lesson.id, class_code="2B", subject="數學",
            teachers_json=json.dumps([absent.id]), from_date=absence.data, from_period=1,
            to_date=absence.data, to_period=1, replaced_teacher_id=absent.id,
            replacement_teacher_id=replacement.id,
        ))
        self.db.commit()

        entries = daily_export_data(self.db, absence.data)
        row = entries[0]["rows"][0]
        self.assertEqual(row["substitute_teacher"], "吳淑君")
        self.assertEqual(row["remark"], "吳淑君代上2B數學")
        workbook = load_workbook(BytesIO(build_daily_xlsx(entries, get_period_times(self.db))))
        self.assertEqual(workbook["劉慧妍"]["G7"].value, "吳淑君代上2B數學")
        self.assertTrue(build_daily_pdf(entries, get_period_times(self.db)).startswith(b"%PDF"))

    def test_post_exam_day_uses_five_periods_in_forms_and_timetable(self):
        teacher = Professor(nom="試後老師", actiu=True)
        version = TimetableVersion(
            effective_from=date(2025, 12, 8), effective_to=date(2025, 12, 31),
            class_filename="試後班別表.xlsx", teacher_filename="試後教師表.xlsx", active=True,
        )
        self.db.add_all([teacher, version])
        self.db.flush()
        self.db.add(TimetableLesson(
            version_id=version.id, weekday=3, period=4, class_code="2B", subject="數學",
            teachers_json=json.dumps([teacher.id]),
        ))
        self.db.commit()
        day = date(2025, 12, 11)

        with self.assertRaises(HTTPException) as invalid_period:
            create_absences_batch(
                AbsenceBatchCreateRequest(items=[AbsenceCreateRequest(
                    professor_id=teacher.id, data=day, periods=[6], reason_type="sick",
                )]),
                self.db,
                SimpleNamespace(username="admin"),
            )
        self.assertEqual(invalid_period.exception.status_code, 400)

        absence = AbsenceCase(
            professor_id=teacher.id, data=day, periods_json="[4]", reason_type="sick",
            status="resolved",
        )
        self.db.add(absence)
        self.db.commit()

        periods = get_period_times(self.db, day)
        self.assertEqual(
            [(item["period"], item["start"], item["end"]) for item in periods],
            [(1, "08:25", "09:15"), (2, "09:15", "10:05"),
             (3, "10:25", "11:15"), (4, "11:15", "12:05"),
             (5, "12:25", "13:00")],
        )
        entries = daily_export_data(self.db, day)
        self.assertEqual([row["period"] for row in entries[0]["rows"]], [1, 2, 3, 4, 5])
        workbook = load_workbook(BytesIO(build_daily_xlsx(entries, periods)))
        sheet = workbook["試後老師"]
        self.assertEqual((sheet["B7"].value, sheet["B13"].value), ("08:25-09:15", "12:25-13:00"))
        self.assertEqual((sheet["A9"].value, sheet["A12"].value), ("一息", "二息"))
        self.assertEqual(sheet.max_row, 13)
        self.assertTrue(build_daily_pdf(entries, periods).startswith(b"%PDF"))
        effective = get_effective_timetable(day, db=self.db)
        self.assertTrue(effective["post_exam"])
        self.assertEqual(effective["period_count"], 5)

    def test_statistics_daily_exports_and_complete_cycle_payload(self):
        absent, swap_teacher, cycle_teacher, cover_teacher = [
            Professor(nom=name, actiu=True) for name in ("甲老師", "乙老師", "丁老師", "丙老師")
        ]
        version = TimetableVersion(
            effective_from=date(2026, 8, 1), effective_to=date(2027, 7, 31),
            class_filename="classes.xls", teacher_filename="teachers.xlsx", active=True,
        )
        self.db.add_all([absent, swap_teacher, cycle_teacher, cover_teacher, version])
        self.db.flush()
        first = TimetableLesson(
            version_id=version.id, weekday=0, period=1, class_code="1A", subject="中文",
            teachers_json=json.dumps([absent.id]),
        )
        second = TimetableLesson(
            version_id=version.id, weekday=0, period=2, class_code="1A", subject="英文",
            teachers_json=json.dumps([swap_teacher.id]),
        )
        third = TimetableLesson(
            version_id=version.id, weekday=0, period=3, class_code="1A", subject="數學",
            teachers_json=json.dumps([cycle_teacher.id]),
        )
        absence = AbsenceCase(
            professor_id=absent.id, data=date(2026, 8, 10), periods_json="[1]",
            reason_type="other", reason_detail="家庭安排", status="resolved",
        )
        self.db.add_all([first, second, third, absence])
        self.db.flush()
        swap = ScheduleAdjustment(absence_case_id=absence.id, kind="three_cycle", status="confirmed", locked=True)
        cover = ScheduleAdjustment(absence_case_id=absence.id, kind="emergency_cover", status="confirmed", locked=True)
        self.db.add_all([swap, cover])
        self.db.flush()
        self.db.add_all([
            ScheduleAdjustmentLeg(
                adjustment_id=swap.id, lesson_id=first.id, class_code="1A", subject="中文",
                teachers_json=json.dumps([absent.id]), from_date=date(2026, 8, 10), from_period=1,
                to_date=date(2026, 8, 10), to_period=3,
            ),
            ScheduleAdjustmentLeg(
                adjustment_id=swap.id, lesson_id=second.id, class_code="1A", subject="英文",
                teachers_json=json.dumps([swap_teacher.id]), from_date=date(2026, 8, 10), from_period=2,
                to_date=date(2026, 8, 10), to_period=1,
            ),
            ScheduleAdjustmentLeg(
                adjustment_id=swap.id, lesson_id=third.id, class_code="1A", subject="數學",
                teachers_json=json.dumps([cycle_teacher.id]), from_date=date(2026, 8, 10), from_period=3,
                to_date=date(2026, 8, 10), to_period=2,
            ),
            ScheduleAdjustmentLeg(
                adjustment_id=cover.id, lesson_id=first.id, class_code="1A", subject="中文",
                teachers_json=json.dumps([absent.id]), from_date=date(2026, 9, 1), from_period=3,
                to_date=date(2026, 9, 1), to_period=3, replaced_teacher_id=absent.id,
                replacement_teacher_id=cover_teacher.id,
            ),
        ])
        self.db.commit()

        statistics = teacher_statistics(
            date(2026, 8, 1), date(2027, 7, 31), self.db, SimpleNamespace(username="admin")
        )
        totals = {row["name"]: row["total"] for row in statistics["teachers"]}
        self.assertEqual(totals, {"丁老師": 1, "丙老師": 1, "乙老師": 1, "甲老師": 0})

        effective = get_effective_timetable(date(2026, 8, 10), db=self.db)
        cycle = next(item for item in effective["adjustments"] if item["id"] == swap.id)
        self.assertEqual(len(cycle["legs"]), 3)

        entries = daily_export_data(self.db, date(2026, 8, 10))
        self.assertEqual(entries[0]["reason_type"], "other")
        self.assertEqual(entries[0]["reason_detail"], "家庭安排")
        self.assertEqual(entries[0]["reason_label"], "其他：家庭安排")
        self.assertEqual(entries[0]["rows"][0]["subject"], "中文")
        self.assertEqual(entries[0]["rows"][0]["substitute_teacher"], "乙老師")
        self.assertEqual(entries[0]["rows"][0]["remark"], "乙老師調上1A英文，第3節甲老師上1A中文")
        periods = get_period_times(self.db)
        workbook = load_workbook(BytesIO(build_daily_xlsx(entries, periods)))
        self.assertEqual(workbook.sheetnames, ["甲老師"])
        self.assertEqual(workbook["甲老師"]["A1"].value, "2026 年")
        self.assertEqual(workbook["甲老師"]["D4"].value, "原因：其他：家庭安排")
        self.assertEqual(workbook["甲老師"]["D5"].value, "原科目")
        self.assertEqual(workbook["甲老師"]["E5"].value, "代課老師")
        self.assertEqual(workbook["甲老師"]["E7"].value, "乙老師")
        self.assertEqual(workbook["甲老師"]["G7"].value, "乙老師調上1A英文，第3節甲老師上1A中文")
        pdf_entries = [deepcopy(entries[0]) for _ in range(3)]
        for index, entry in enumerate(pdf_entries, 1):
            entry["teacher_id"] = index
            entry["teacher_name"] = f"第{index}位老師全名"
        for count, expected_pages in ((1, 1), (2, 1), (3, 2)):
            pdf = build_daily_pdf(pdf_entries[:count], periods)
            self.assertTrue(pdf.startswith(b"%PDF"))
            self.assertEqual(len(re.findall(rb"/Type\s*/Page(?!s)", pdf)), expected_pages)
        with patch("routes.rescheduling.daily_export_data", return_value=pdf_entries):
            response = export_daily_pdf(
                date(2026, 8, 10), self.db, SimpleNamespace(username="admin")
            )
        self.assertEqual(response.media_type, "application/pdf")
        self.assertTrue(response.headers["content-disposition"].endswith('.pdf"'))
        legs = (
            self.db.query(ScheduleAdjustmentLeg)
            .filter_by(adjustment_id=swap.id)
            .order_by(ScheduleAdjustmentLeg.id)
            .all()
        )
        swap.kind = "direct_swap"
        legs[0].to_period = 2
        self.db.delete(legs[2])
        self.db.commit()
        direct = daily_export_data(self.db, date(2026, 8, 10))
        self.assertEqual(
            direct[0]["rows"][0]["remark"],
            "乙老師調上1A英文，第2節甲老師上1A中文",
        )
        legs[0].to_date = date(2026, 8, 11)
        self.db.commit()
        cross_day = daily_export_data(self.db, date(2026, 8, 10))
        self.assertEqual(
            cross_day[0]["rows"][0]["remark"],
            "乙老師調上1A英文，08/11第2節甲老師上1A中文",
        )
        changed_periods = [{**item, "start": "08:20"} if item["period"] == 1 else item for item in periods]
        save_period_times(self.db, changed_periods)
        self.assertEqual(get_period_times(self.db)[0]["start"], "08:20")
        with self.assertRaises(ValueError):
            validate_period_times([{**item, "start": "08:30"} if item["period"] == 2 else item for item in periods])


if __name__ == "__main__":
    unittest.main()
