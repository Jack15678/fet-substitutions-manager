"""Daily rescheduling/substitution forms shared by XLSX and PDF exports."""
from io import BytesIO
import json
import os
from pathlib import Path
import re
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import (
    Frame,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from models import AbsenceCase, Professor, ScheduleAdjustment, ScheduleAdjustmentLeg, TimetableLesson
from repositories import ConfiguracioRepository
from rescheduling_service import effective_occurrences, is_post_exam_version, version_for_date


DEFAULT_PERIOD_TIMES = [
    {"period": 1, "start": "08:25", "end": "09:00"},
    {"period": 2, "start": "09:00", "end": "09:35"},
    {"period": 3, "start": "09:35", "end": "10:10"},
    {"period": 4, "start": "10:25", "end": "11:00"},
    {"period": 5, "start": "11:00", "end": "11:35"},
    {"period": 6, "start": "11:35", "end": "12:10"},
    {"period": 7, "start": "13:00", "end": "13:35"},
    {"period": 8, "start": "13:35", "end": "14:10"},
    {"period": 9, "start": "14:25", "end": "15:00"},
]
POST_EXAM_PERIOD_TIMES = [
    {"period": 1, "start": "08:25", "end": "09:15"},
    {"period": 2, "start": "09:15", "end": "10:05"},
    {"period": 3, "start": "10:25", "end": "11:15"},
    {"period": 4, "start": "11:15", "end": "12:05"},
    {"period": 5, "start": "12:25", "end": "13:00"},
]
WEEKDAYS = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
ABSENCE_REASONS = {
    "sick": "病假",
    "follow_up": "覆診",
    "team_training": "帶隊訓練",
    "other": "其他",
    # Legacy values retained for existing records.
    "personal": "事假",
    "official": "公假",
    "training": "培訓",
}


def _pdf_font() -> str:
    name = "DailyCJK"
    try:
        pdfmetrics.getFont(name)
        return name
    except KeyError:
        pass
    windows = Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts"
    candidates = [
        os.getenv("CJK_FONT_PATH"),
        windows / "msjh.ttc",
        windows / "simhei.ttf",
        windows / "NotoSansJP-Regular.ttf",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).is_file():
            pdfmetrics.registerFont(TTFont(name, str(candidate), subfontIndex=0))
            return name
    raise RuntimeError("找不到可嵌入 PDF 的中文字型，請安裝文泉驛正黑或設定 CJK_FONT_PATH")


def validate_period_times(items: list[dict], expected_periods: int = 9) -> list[dict]:
    normalized = sorted(
        ({"period": int(item["period"]), "start": item["start"], "end": item["end"]} for item in items),
        key=lambda item: item["period"],
    )
    if [item["period"] for item in normalized] != list(range(1, expected_periods + 1)):
        raise ValueError(f"必須完整設定第 1 至第 {expected_periods} 節")

    previous_end = -1
    for item in normalized:
        if not re.fullmatch(r"(?:[01]\d|2[0-3]):[0-5]\d", item["start"]) or not re.fullmatch(
            r"(?:[01]\d|2[0-3]):[0-5]\d", item["end"]
        ):
            raise ValueError("節次時間必須使用 HH:MM 格式")
        start = int(item["start"][:2]) * 60 + int(item["start"][3:])
        end = int(item["end"][:2]) * 60 + int(item["end"][3:])
        if start >= end:
            raise ValueError("每節結束時間必須晚於開始時間")
        if start < previous_end:
            raise ValueError("節次時間不可重疊或倒序")
        previous_end = end
    return normalized


def get_period_times(db, day=None) -> list[dict]:
    version = version_for_date(db, day) if day is not None else None
    if version and is_post_exam_version(version):
        return [dict(item) for item in POST_EXAM_PERIOD_TIMES]
    raw = ConfiguracioRepository.get(db, "rescheduling_period_times")
    if raw:
        try:
            return validate_period_times(json.loads(raw))
        except (KeyError, TypeError, ValueError, json.JSONDecodeError):
            pass
    return [dict(item) for item in DEFAULT_PERIOD_TIMES]


def save_period_times(db, items: list[dict]) -> list[dict]:
    normalized = validate_period_times(items)
    ConfiguracioRepository.set(
        db,
        "rescheduling_period_times",
        json.dumps(normalized, ensure_ascii=False),
        tipus="json",
        descripcio="調課／代課表第 1 至第 9 節時間",
    )
    return normalized


def _unique(values) -> str:
    return "、".join(dict.fromkeys(value for value in values if value))


def _absence_reason(reason_type: str | None, reason_detail: str | None) -> str:
    if not reason_type:
        return "未填寫"
    label = ABSENCE_REASONS.get(reason_type, "未填寫")
    return f"{label}：{reason_detail}" if reason_type == "other" and reason_detail else label


def daily_export_data(db, day) -> list[dict]:
    cases = (
        db.query(AbsenceCase)
        .filter(AbsenceCase.data == day, AbsenceCase.status == "resolved")
        .order_by(AbsenceCase.professor_id, AbsenceCase.id)
        .all()
    )
    if not cases:
        return []

    names = {row.id: row.nom for row in db.query(Professor).all()}
    cases_by_teacher = {case.professor_id: case for case in cases}
    periods_by_teacher: dict[int, set[int]] = {}
    for case in cases:
        periods_by_teacher.setdefault(case.professor_id, set()).update(
            int(value) for value in json.loads(case.periods_json or "[]")
        )

    version = version_for_date(db, day)
    period_count = len(get_period_times(db, day))
    base_lessons = db.query(TimetableLesson).filter_by(
        version_id=version.id, weekday=day.weekday()
    ).all() if version else []
    effective = effective_occurrences(db, day, day)
    confirmed_legs = (
        db.query(ScheduleAdjustmentLeg, ScheduleAdjustment)
        .join(ScheduleAdjustment, ScheduleAdjustmentLeg.adjustment_id == ScheduleAdjustment.id)
        .filter(
            ScheduleAdjustment.status == "confirmed",
            ScheduleAdjustment.absence_case_id.in_([case.id for case in cases]),
        )
        .order_by(ScheduleAdjustment.id, ScheduleAdjustmentLeg.id)
        .all()
    )
    result = []

    for teacher_id, absent_periods in periods_by_teacher.items():
        rows = []
        for period in range(1, period_count + 1):
            originals = [
                lesson for lesson in base_lessons
                if lesson.period == period and teacher_id in json.loads(lesson.teachers_json or "[]")
            ] if period in absent_periods else []
            covered_moves = [
                leg for leg, adjustment in confirmed_legs
                if adjustment.kind in {"emergency_cover", "co_teacher_solo"}
                and leg.from_date == day and leg.from_period == period
                and leg.replaced_teacher_id == teacher_id
            ] if period in absent_periods else []
            covered_slots = {(leg.lesson_id, leg.from_date, leg.from_period) for leg in covered_moves}
            scheduled = [*originals, *covered_moves]
            original_teacher_ids = {
                int(value) for lesson in scheduled for value in json.loads(lesson.teachers_json or "[]")
            }
            class_codes = [lesson.class_code for lesson in scheduled]
            actual = [
                occurrence for occurrence in effective
                if occurrence["period"] == period and occurrence["class_code"] in class_codes
            ]
            actual_teacher_ids = [
                value for occurrence in actual for value in occurrence["teachers"] if value != teacher_id
            ]
            related_adjustment_ids = {
                adjustment.id for leg, adjustment in confirmed_legs
                if adjustment.kind not in {"emergency_cover", "co_teacher_solo"} and leg.to_period == period
                and leg.to_date == day and leg.class_code in class_codes
            }
            remark_teacher_ids = {teacher_id, *actual_teacher_ids}
            remarks = []
            original_remarks = []
            for leg, adjustment in confirmed_legs:
                if adjustment.kind == "emergency_cover":
                    if (leg.to_period == period and leg.to_date == day
                            and leg.class_code in class_codes and leg.replacement_teacher_id):
                        replacement_id = int(leg.replacement_teacher_id)
                        remarks.append(
                            f"{names.get(replacement_id, str(replacement_id))}"
                            f"代上{leg.class_code}{leg.subject}"
                        )
                    continue
                if adjustment.id not in related_adjustment_ids:
                    continue
                for value in json.loads(leg.teachers_json or "[]"):
                    teacher = int(value)
                    if teacher == teacher_id:
                        if (leg.lesson_id, leg.to_date, leg.to_period) in covered_slots:
                            continue
                        original_remarks.append(
                            (f"{leg.to_date:%d/%m}" if leg.to_date != day else "") +
                            f"第{leg.to_period}節"
                            f"{names.get(teacher, str(teacher))}上{leg.class_code}{leg.subject}"
                        )
                    elif teacher in remark_teacher_ids:
                        remarks.append(
                            f"{names.get(teacher, str(teacher))}"
                            f"調上{leg.class_code}{leg.subject}"
                        )
            rows.append({
                "period": period,
                "class_code": _unique(class_codes),
                "subject": _unique(lesson.subject for lesson in scheduled),
                "substitute_teacher": _unique(
                    names.get(value, str(value)) for value in actual_teacher_ids
                    if value not in original_teacher_ids
                ),
                "remark": "，".join(dict.fromkeys(remarks + original_remarks)),
            })
        absence = cases_by_teacher[teacher_id]
        result.append({
            "date": day,
            "weekday": WEEKDAYS[day.weekday()],
            "teacher_id": teacher_id,
            "teacher_name": names.get(teacher_id, str(teacher_id)),
            "reason_type": absence.reason_type,
            "reason_detail": absence.reason_detail,
            "reason_label": _absence_reason(absence.reason_type, absence.reason_detail),
            "rows": rows,
        })
    return result


def _safe_sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", name).strip()[:31] or "教師"
    candidate = base
    counter = 2
    while candidate in used:
        suffix = f"-{counter}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def _export_times(period_times: list[dict]) -> dict[int, dict]:
    if len(period_times) not in {5, 9}:
        raise ValueError("匯出節次必須完整設定為 5 節或 9 節")
    return {item["period"]: item for item in validate_period_times(period_times, len(period_times))}


def build_daily_xlsx(entries: list[dict], period_times: list[dict]) -> bytes:
    times = _export_times(period_times)
    break_labels = {2: "一息", 4: "二息"} if len(times) == 5 else {3: "一息", 6: "午膳", 8: "二息"}
    row_map = {}
    break_rows = []
    next_row = 7
    for period in times:
        row_map[period] = next_row
        next_row += 1
        if period in break_labels:
            break_rows.append((next_row, break_labels[period]))
            next_row += 1
    last_row = next_row - 1
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names: set[str] = set()
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    gray = PatternFill("solid", fgColor="D9D9D9")
    title_font = Font(name="PMingLiU", size=16, bold=True)
    body_font = Font(name="PMingLiU", size=11)

    for entry in entries:
        sheet = workbook.create_sheet(_safe_sheet_name(entry["teacher_name"], used_names))
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_margins.left = sheet.page_margins.right = 0.25
        sheet.page_margins.top = sheet.page_margins.bottom = 0.35
        widths = [6, 16, 14, 18, 19, 14, 25]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

        sheet.merge_cells("A1:G1")
        sheet["A1"] = f"{entry['date'].year} 年"
        sheet.merge_cells("A2:G2")
        sheet["A2"] = "調課／代課表（全日制）"
        sheet.merge_cells("A3:G3")
        day = entry["date"]
        sheet["A3"] = f"代課日期：{day.year} 年 {day.month} 月 {day.day} 日（{entry['weekday']}）"
        sheet.merge_cells("A4:C4")
        sheet["A4"] = f"請假教師：{entry['teacher_name']}"
        sheet.merge_cells("D4:G4")
        sheet["D4"] = f"原因：{entry['reason_label']}"
        sheet.merge_cells("A5:B5")
        headers = {"A5": "節次", "C5": "班別", "D5": "原科目", "E5": "代課老師", "F5": "簽署", "G5": "附註"}
        for cell, value in headers.items():
            sheet[cell] = value

        sheet.merge_cells("A6:B6")
        sheet["A6"] = "班主任課"
        for row, label in break_rows:
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            sheet.cell(row, 1, label)
            for column in range(1, 8):
                sheet.cell(row, column).fill = gray

        for item in entry["rows"]:
            row = row_map[item["period"]]
            timing = times[item["period"]]
            values = [item["period"], f"{timing['start']}-{timing['end']}", item["class_code"], item["subject"], item["substitute_teacher"], "", item["remark"]]
            for column, value in enumerate(values, start=1):
                sheet.cell(row, column, value)

        for row in range(1, last_row + 1):
            sheet.row_dimensions[row].height = 24 if row > 4 else 22
            for column in range(1, 8):
                cell = sheet.cell(row, column)
                cell.font = title_font if row in (1, 2) else body_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in range(5, last_row + 1):
            for column in range(1, 8):
                sheet.cell(row, column).border = border
        sheet["A4"].alignment = Alignment(horizontal="left", vertical="center")
        sheet["D4"].alignment = Alignment(horizontal="left", vertical="center")
        sheet.print_area = f"A1:G{last_row}"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_daily_pdf(entries: list[dict], period_times: list[dict]) -> bytes:
    times = _export_times(period_times)
    break_labels = {2: "一息", 4: "二息"} if len(times) == 5 else {3: "一息", 6: "午膳", 8: "二息"}
    font = _pdf_font()
    output = BytesIO()
    canvas = Canvas(output, pagesize=A4)
    page_width, page_height = A4
    margin = 5 * mm
    half_height = (page_height - 2 * margin) / 2
    frame_width = page_width - 2 * margin
    frame_padding = 2 * mm
    inner_width = frame_width - 2 * frame_padding
    half_inner_height = half_height - 2 * frame_padding
    title = ParagraphStyle("daily-title", fontName=font, fontSize=15.5, leading=17, alignment=TA_CENTER)
    info = ParagraphStyle("daily-info", fontName=font, fontSize=12, leading=14, alignment=TA_LEFT)
    note = ParagraphStyle(
        "daily-note", fontName=font, fontSize=12, leading=12, alignment=TA_CENTER, wordWrap="CJK"
    )
    half_slot = 0
    for entry in entries:
        day = entry["date"]
        form = [
            Paragraph(f"{day.year} 年　調課／代課表（全日制）", title),
            Spacer(1, 1 * mm),
            Paragraph(
                f"代課日期：{day.year} 年 {day.month} 月 {day.day} 日（{entry['weekday']}）", info,
            ),
            Paragraph(
                f"請假教師：{escape(entry['teacher_name'])}　　原因：{escape(entry['reason_label'])}", info,
            ),
            Spacer(1, 1.5 * mm),
        ]
        rows_by_period = {row["period"]: row for row in entry["rows"]}
        table_data = [
            ["節次", "", "班別", "原科目", "代課老師", "簽署", "附註"],
            ["班主任課", "", "", "", "", "", ""],
        ]
        row_heights = [7.5 * mm, 7.5 * mm]
        spans = [("SPAN", (0, 0), (1, 0)), ("SPAN", (0, 1), (1, 1))]
        for period in times:
            row = rows_by_period[period]
            timing = times[period]
            remark = Paragraph(escape(row["remark"]), note) if row["remark"] else ""
            table_data.append([
                period,
                f"{timing['start']}-{timing['end']}",
                row["class_code"],
                row["subject"],
                row["substitute_teacher"],
                "",
                remark,
            ])
            row_heights.append(
                max(7 * mm, remark.wrap(88 * mm - 12, page_height)[1] + 2) if remark else 7 * mm
            )
            if period in break_labels:
                label = break_labels[period]
                table_data.append([label, "", "", "", "", "", ""])
                row_heights.append(6 * mm)
                spans.append(("SPAN", (0, len(table_data) - 1), (-1, len(table_data) - 1)))

        table = Table(
            table_data,
            colWidths=[12 * mm, 28 * mm, 14 * mm, 18 * mm, 24 * mm, 12 * mm, 88 * mm],
            rowHeights=row_heights,
        )
        style = [
            ("FONTNAME", (0, 0), (-1, -1), font),
            ("FONTSIZE", (0, 0), (-1, -1), 12),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.75, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
        ]
        for command in spans:
            style.append(command)
            if command[1][1] >= 2:
                style.append(("BACKGROUND", command[1], command[2], colors.HexColor("#D9D9D9")))
        table.setStyle(TableStyle(style))
        form.append(table)
        full_page = sum(item.wrap(inner_width, page_height)[1] for item in form) > half_inner_height
        if full_page and half_slot:
            canvas.showPage()
            half_slot = 0
        frame_height = page_height - 2 * margin if full_page else half_height
        frame = Frame(
            margin,
            margin if full_page or half_slot else margin + half_height,
            frame_width,
            frame_height,
            leftPadding=frame_padding,
            rightPadding=frame_padding,
            topPadding=frame_padding,
            bottomPadding=frame_padding,
        )
        frame.addFromList(form, canvas)
        if form:
            raise RuntimeError("每日 PDF 表格超出頁面範圍")
        if full_page or half_slot:
            canvas.showPage()
            half_slot = 0
        else:
            half_slot = 1
    if half_slot:
        canvas.showPage()
    canvas.save()
    return output.getvalue()
