"""Excel import and deterministic lesson-rescheduling rules.

The school workbooks are the immutable base timetable. Confirmed adjustments are
stored separately and overlaid whenever an effective timetable is requested.
"""
from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
import hashlib
from itertools import permutations
import json
import re
from typing import Iterable

import openpyxl
import xlrd
from sqlalchemy import or_
from sqlalchemy.orm import Session

from models import (
    AbsenceCase,
    Configuracio,
    Professor,
    ProfessorBaixa,
    ScheduleAdjustment,
    ScheduleAdjustmentLeg,
    SchoolClosure,
    TimetableLesson,
    TimetableTeacherSlot,
    TimetableVersion,
)


NAME_ALIASES = {"李曉琳": "李晓琳"}
EMPTY_MARKERS = {"", "-", "--", "—", "/", "x", "X", "-x-"}
IGNORED_CLASS_ACTIVITIES = {"康樂活動(13:00-14:30)"}
WEEKDAYS = {f"星期{name}": index for index, name in enumerate("一二三四五")}
TIME_RANGE = re.compile(r"(\d{1,2}):(\d{2})\s*[-–—]\s*(\d{1,2}):(\d{2})")
MAX_CYCLE_LESSONS_KEY = "rescheduling_max_cycle_lessons"
CORE_CONSECUTIVE_SUBJECTS = {"中文", "中國語文", "英文", "英語", "數學"}
REPAIRABLE_SWAP_KINDS = {"direct_swap", "three_cycle", "manual_swap", "manual_three_cycle"}


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ").strip()
    return re.sub(r"\s+", " ", text)


def normalize_name(value) -> str:
    value = clean(value).replace(" ", "")
    return NAME_ALIASES.get(value, value)


def normalize_class(value) -> str:
    return clean(value).replace(" ", "").upper()


def normalize_subject(value) -> str:
    value = clean(value)
    if value.endswith("課") and len(value) > 1:
        value = value[:-1]
    return value


def adjacent_teaching_count(occurrences: list[dict], teacher_id: int,
                            target_date: date, target_period: int) -> int:
    """Count adjacent formal periods in which the teacher is already teaching."""
    return sum(any(
        occurrence["date"] == target_date
        and occurrence["period"] == period
        and teacher_id in occurrence["teachers"]
        for occurrence in occurrences
    ) for period in (target_period - 1, target_period + 1) if 1 <= period <= 9)


def _core_consecutive_pairs(occurrences: list[dict]) -> set[tuple[str, str]]:
    by_slot: dict[tuple, set[str]] = {}
    for occurrence in occurrences:
        subject = normalize_subject(occurrence["subject"])
        if occurrence["lesson_id"] is None or subject not in CORE_CONSECUTIVE_SUBJECTS:
            continue
        for teacher_id in occurrence["teachers"]:
            key = (int(teacher_id), occurrence["date"], occurrence["class_code"],
                   subject, int(occurrence["period"]))
            by_slot.setdefault(key, set()).add(occurrence["occurrence_id"])
    pairs = set()
    for key, occurrence_ids in by_slot.items():
        teacher_id, day, class_code, subject, period = key
        next_ids = by_slot.get((teacher_id, day, class_code, subject, period + 1), set())
        pairs.update(tuple(sorted((left, right))) for left in occurrence_ids for right in next_ids)
    return pairs


def _broken_core_consecutive_count(legs: list[dict], occurrence_slots: dict[str, tuple[date, int]],
                                   pairs: set[tuple[str, str]]) -> int:
    destinations = {
        leg["occurrence_id"]: (date.fromisoformat(leg["to_date"]), int(leg["to_period"]))
        for leg in legs
    }
    broken = 0
    for left, right in pairs:
        if left not in destinations and right not in destinations:
            continue
        left_slot = destinations.get(left, occurrence_slots[left])
        right_slot = destinations.get(right, occurrence_slots[right])
        broken += left_slot[0] != right_slot[0] or abs(left_slot[1] - right_slot[1]) != 1
    return broken


def _is_real(value: str) -> bool:
    return clean(value) not in EMPTY_MARKERS


def _time_key(value) -> tuple[int, int] | None:
    match = TIME_RANGE.search(clean(value))
    if not match:
        return None
    start_hour, start_minute, end_hour, end_minute = map(int, match.groups())
    return start_hour * 60 + start_minute, end_hour * 60 + end_minute


def teacher_names_from_workbook(content: bytes) -> list[str]:
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    names = []
    for sheet in workbook.worksheets:
        teacher = ""
        for row in sheet.iter_rows(min_row=1, max_row=min(10, sheet.max_row), max_col=min(12, sheet.max_column)):
            for index, cell in enumerate(row[:-1]):
                if clean(cell.value).replace(" ", "") in {"教師:", "教師："}:
                    teacher = normalize_name(row[index + 1].value)
                    break
            if teacher:
                break
        names.append(teacher or normalize_name(sheet.title))
    workbook.close()
    return sorted({name for name in names if name})


def _openpyxl_weekday_columns(sheet) -> dict[int, int]:
    for row_number in range(1, min(12, sheet.max_row) + 1):
        found = {
            WEEKDAYS[clean(sheet.cell(row=row_number, column=column).value).replace(" ", "")]: column
            for column in range(1, min(15, sheet.max_column) + 1)
            if clean(sheet.cell(row=row_number, column=column).value).replace(" ", "") in WEEKDAYS
        }
        if len(found) == 5:
            return found
    raise ValueError(f"教師表工作表 {sheet.title} 找不到星期一至五標題")


def parse_teacher_workbook(content: bytes, period_times: list[tuple[int, int]],
                           names: list[str] | None = None) -> tuple[list[dict], list[str]]:
    """Read only the nine teaching-period rows; duty/recess rows are ignored."""
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    slots: list[dict] = []
    names = names or teacher_names_from_workbook(content)
    for sheet in workbook.worksheets:
        weekday_columns = _openpyxl_weekday_columns(sheet)
        teacher = next((name for name in names if name == normalize_name(sheet.title)), "")
        if not teacher:
            for row in sheet.iter_rows(min_row=1, max_row=min(10, sheet.max_row), max_col=min(12, sheet.max_column)):
                for index, cell in enumerate(row[:-1]):
                    if clean(cell.value).replace(" ", "") in {"教師:", "教師："}:
                        teacher = normalize_name(row[index + 1].value)
                        break
                if teacher:
                    break
        if not teacher:
            teacher = normalize_name(sheet.title)
        rows_by_time = {}
        for row_number in range(1, sheet.max_row + 1):
            for column in range(1, min(weekday_columns.values())):
                key = _time_key(sheet.cell(row=row_number, column=column).value)
                if key:
                    rows_by_time[key] = row_number
                    break
        missing = [key for key in period_times if key not in rows_by_time]
        if missing:
            raise ValueError(f"教師表工作表 {sheet.title} 缺少 {len(missing)} 個正式教學節次")
        for period, key in enumerate(period_times, start=1):
            class_row = rows_by_time[key]
            subject_row = class_row + 1
            for weekday, column in weekday_columns.items():
                class_codes = [normalize_class(value) for value in re.split(
                    r"[、,，/＋+;；\n]+", clean(sheet.cell(row=class_row, column=column).value)
                ) if normalize_class(value)]
                if class_codes:
                    prefix = re.match(r"(\d+)", class_codes[0])
                    class_codes = [
                        f"{prefix.group(1)}{value}" if prefix and value.isalpha() else value
                        for value in class_codes
                    ]
                subjects = [normalize_subject(value) for value in re.split(
                    r"[、,，/＋+;；\n]+", clean(sheet.cell(row=subject_row, column=column).value)
                ) if normalize_subject(value)]
                if not (class_codes or subjects):
                    continue
                class_codes = class_codes or [""]
                subjects = subjects or [""]
                for index in range(max(len(class_codes), len(subjects))):
                    slots.append({
                        "teacher": teacher,
                        "weekday": weekday,
                        "period": period,
                        "class_code": class_codes[min(index, len(class_codes) - 1)],
                        "subject": subjects[min(index, len(subjects) - 1)],
                    })
    workbook.close()
    return slots, names


def _teachers_from_cell(value, known_names: Iterable[str]) -> list[str]:
    raw = clean(value).replace(" ", "")
    if not _is_real(raw):
        return []
    found = [name for name in sorted(set(known_names), key=len, reverse=True)
             if name and (name in raw or any(alias in raw for alias, target in NAME_ALIASES.items()
                                               if target == name))]
    if found:
        # A short name can be contained in a longer one (e.g. 李雅 / 李雅妍).
        return [name for name in dict.fromkeys(found)
                if not any(name != other and name in other for other in found)]
    return list(dict.fromkeys(
        normalize_name(part)
        for part in re.split(r"[、,，/＋+;；\n]+", clean(value))
        if normalize_name(part)
    ))


def _contains_known_teacher(value, known_names: Iterable[str]) -> bool:
    raw = clean(value).replace(" ", "")
    return any(name in raw or any(alias in raw for alias, target in NAME_ALIASES.items() if target == name)
               for name in known_names if name)


def _single_sheet(content: bytes, label: str):
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    if len(workbook.worksheets) != 1:
        workbook.close()
        raise ValueError(f"{label}必須只有一張工作表")
    return workbook, workbook.active


def _split_classes(value) -> list[str]:
    classes = [normalize_class(part) for part in re.split(
        r"[、,，/＋+;；\n]+", clean(value)
    ) if normalize_class(part)]
    if classes:
        prefix = re.match(r"(\d+)", classes[0])
        classes = [f"{prefix.group(1)}{item}" if prefix and item.isalpha() else item
                   for item in classes]
    return classes


def _split_subjects(value) -> list[str]:
    return [normalize_subject(part) for part in re.split(
        r"[、,，/＋+;；\n]+", clean(value)
    ) if normalize_subject(part)]


def _matrix_header_row(sheet, label: str) -> int:
    for row in range(1, min(10, sheet.max_row) + 1):
        if clean(sheet.cell(row=row, column=1).value).replace(" ", "") in {"班別", "教師"}:
            return row
    raise ValueError(f"{label}找不到橫排標題")


def _matrix_period_rows(sheet, header_row: int, label: str,
                        end_row: int | None = None) -> list[tuple[tuple[int, int], int]]:
    rows = []
    for row in range(header_row + 1, (end_row or sheet.max_row) + 1):
        key = _time_key(sheet.cell(row=row, column=1).value)
        if key and 25 <= key[1] - key[0] <= 55:
            rows.append((key, row))
    if len(rows) != 5:
        raise ValueError(f"{label}辨識到 {len(rows)} 個正式教學節次，預期為 5 個")
    return rows


def parse_post_exam_class_workbook(
    content: bytes, known_names: Iterable[str]
) -> tuple[list[dict], list[tuple[int, int]]]:
    workbook, sheet = _single_sheet(content, "試後班別表")
    header_rows = [row for row in range(1, sheet.max_row + 1)
                   if clean(sheet.cell(row=row, column=1).value).replace(" ", "") == "班別"]
    if not header_rows:
        workbook.close()
        raise ValueError("試後班別表找不到橫排標題")
    lessons: list[dict] = []
    class_codes: set[str] = set()
    expected_times: list[tuple[int, int]] | None = None
    for index, header_row in enumerate(header_rows):
        columns = [
            (column, normalize_class(sheet.cell(row=header_row, column=column).value))
            for column in range(2, sheet.max_column + 1)
            if normalize_class(sheet.cell(row=header_row, column=column).value)
        ]
        block_codes = {class_code for _, class_code in columns}
        if len(block_codes) != len(columns) or class_codes & block_codes:
            workbook.close()
            raise ValueError("試後班別表有重複班別")
        class_codes.update(block_codes)
        end_row = header_rows[index + 1] - 1 if index + 1 < len(header_rows) else sheet.max_row
        period_rows = _matrix_period_rows(sheet, header_row, "試後班別表", end_row)
        times = [key for key, _ in period_rows]
        if expected_times is None:
            expected_times = times
        elif times != expected_times:
            workbook.close()
            raise ValueError("試後班別表各區塊的上課時間不一致")
        for period, (_, row) in enumerate(period_rows, start=1):
            for column, class_code in columns:
                raw = clean(sheet.cell(row=row, column=column).value)
                teacher_row = ""
                if (row < end_row
                        and not _time_key(sheet.cell(row=row + 1, column=1).value)
                        and _contains_known_teacher(
                            sheet.cell(row=row + 1, column=column).value, known_names
                        )):
                    teacher_row = clean(sheet.cell(row=row + 1, column=column).value)
                teachers = _teachers_from_cell(f"{raw}\n{teacher_row}", known_names)
                subject = raw
                for teacher in teachers:
                    subject = subject.replace(teacher, "")
                    for alias, target in NAME_ALIASES.items():
                        if target == teacher:
                            subject = subject.replace(alias, "")
                subject = normalize_subject(subject.strip(" /、,，-—"))
                if not _is_real(subject):
                    workbook.close()
                    raise ValueError(f"試後班別表 {class_code} 第 {period} 節缺少科目")
                for weekday in range(5):
                    lessons.append({
                        "weekday": weekday,
                        "period": period,
                        "class_code": class_code,
                        "subject": subject,
                        "teachers": teachers,
                    })
    if len(class_codes) != 25:
        workbook.close()
        raise ValueError("試後班別表必須合共列出 25 個不同班別")
    workbook.close()
    return lessons, expected_times or []


def parse_post_exam_teacher_workbook(content: bytes) -> tuple[list[dict], list[str]]:
    workbook, sheet = _single_sheet(content, "試後教師表")
    header_row = _matrix_header_row(sheet, "試後教師表")
    columns = []
    for column in range(2, sheet.max_column + 1):
        teacher = normalize_name(re.sub(
            r"^\d+\s*", "", clean(sheet.cell(row=header_row, column=column).value)
        ))
        if teacher:
            columns.append((column, teacher))
    names = [teacher for _, teacher in columns]
    if not names or len(set(names)) != len(names):
        workbook.close()
        raise ValueError("試後教師表必須橫向列出教師的完整姓名，且不可重複")
    period_rows = _matrix_period_rows(sheet, header_row, "試後教師表")
    slots: list[dict] = []
    for period, (_, row) in enumerate(period_rows, start=1):
        for column, teacher in columns:
            raw = str(sheet.cell(row=row, column=column).value or "")
            if row < sheet.max_row and not _time_key(sheet.cell(row=row + 1, column=1).value):
                class_row = sheet.cell(row=row + 1, column=column).value
                classes = _split_classes(class_row)
                if classes and all(re.fullmatch(r"\d+[A-Z]", code) for code in classes):
                    raw = f"{raw}\n{class_row}"
            lines = [clean(line) for line in re.split(
                r"[\r\n]+", raw
            ) if clean(line)]
            if not lines:
                continue
            class_codes = _split_classes(lines[-1])
            if not class_codes or not all(re.fullmatch(r"\d+[A-Z]", code) for code in class_codes):
                class_codes, subjects = [], _split_subjects("、".join(lines))
            else:
                subjects = _split_subjects("、".join(lines[:-1]))
            class_codes = class_codes or [""]
            subjects = subjects or [""]
            for index in range(max(len(class_codes), len(subjects))):
                for weekday in range(5):
                    slots.append({
                        "teacher": teacher,
                        "weekday": weekday,
                        "period": period,
                        "class_code": class_codes[min(index, len(class_codes) - 1)],
                        "subject": subjects[min(index, len(subjects) - 1)],
                    })
    workbook.close()
    return slots, sorted(names)


def _xlrd_weekday_columns(sheet) -> dict[int, int]:
    for row in range(min(12, sheet.nrows)):
        found = {
            WEEKDAYS[clean(sheet.cell_value(row, column)).replace(" ", "")]: column
            for column in range(min(15, sheet.ncols))
            if clean(sheet.cell_value(row, column)).replace(" ", "") in WEEKDAYS
        }
        if len(found) == 5:
            return found
    raise ValueError(f"班別表工作表 {sheet.name} 找不到星期一至五標題")


def parse_class_workbook(content: bytes, known_names: Iterable[str]) -> tuple[list[dict], list[tuple[int, int]]]:
    workbook = xlrd.open_workbook(file_contents=content)
    lessons: list[dict] = []
    expected_times: list[tuple[int, int]] | None = None
    for sheet in workbook.sheets():
        class_code = normalize_class(sheet.name)
        if not class_code:
            continue
        weekday_columns = _xlrd_weekday_columns(sheet)
        pairs = []
        for subject_row in range(sheet.nrows - 1):
            key = next((
                _time_key(sheet.cell_value(subject_row, column))
                for column in range(min(weekday_columns.values()))
                if _time_key(sheet.cell_value(subject_row, column))
            ), None)
            if not key or not 25 <= key[1] - key[0] <= 45:
                continue
            teacher_row = subject_row + 1
            if any(_contains_known_teacher(sheet.cell_value(teacher_row, column), known_names)
                   for column in weekday_columns.values()):
                pairs.append((key, subject_row, teacher_row))
        if len(pairs) != 9:
            raise ValueError(f"班別表工作表 {sheet.name} 辨識到 {len(pairs)} 個教學節次，預期為 9 個")
        period_times = [key for key, _, _ in pairs]
        if expected_times is None:
            expected_times = period_times
        elif period_times != expected_times:
            raise ValueError(f"班別表工作表 {sheet.name} 的上課時間與其他班別不一致")
        for period, (_, subject_row, teacher_row) in enumerate(pairs, start=1):
            for weekday, column in weekday_columns.items():
                subject_raw = clean(sheet.cell_value(subject_row, column))
                subject = normalize_subject(subject_raw)
                if not _is_real(subject):
                    continue
                if subject.replace(" ", "") in IGNORED_CLASS_ACTIVITIES:
                    continue
                teachers = _teachers_from_cell(
                    f"{clean(sheet.cell_value(teacher_row, column))}\n{subject_raw}", known_names
                )
                for teacher in teachers:
                    subject = subject.replace(teacher, "")
                    for alias, target in NAME_ALIASES.items():
                        if target == teacher:
                            subject = subject.replace(alias, "")
                subject = subject.strip(" /、,，-—")
                lessons.append({
                    "weekday": weekday,
                    "period": period,
                    "class_code": class_code,
                    "subject": subject,
                    "teachers": teachers,
                })
    return lessons, expected_times or []


def build_import_preview(class_content: bytes, teacher_content: bytes, post_exam: bool = False) -> dict:
    if post_exam:
        teacher_slots, teacher_names = parse_post_exam_teacher_workbook(teacher_content)
        lessons, period_times = parse_post_exam_class_workbook(class_content, teacher_names)
    else:
        teacher_names = teacher_names_from_workbook(teacher_content)
        lessons, period_times = parse_class_workbook(class_content, teacher_names)
        teacher_slots, teacher_names = parse_teacher_workbook(teacher_content, period_times, teacher_names)
    warnings: list[str] = []
    issues: list[dict] = []
    known = set(teacher_names)
    unknown = sorted({name for lesson in lessons for name in lesson["teachers"] if name not in known})
    empty_lessons = [lesson for lesson in lessons if not lesson["teachers"]]
    empty_teacher = len(empty_lessons)
    if unknown:
        warnings.append("班別表內有未能對應教師表的姓名：" + "、".join(unknown))
        for lesson in lessons:
            for teacher in [name for name in lesson["teachers"] if name not in known]:
                issues.append({
                    "code": "unknown_teacher", "severity": "error",
                    "weekday": lesson["weekday"], "period": lesson["period"],
                    "class_code": lesson["class_code"], "subject": lesson["subject"],
                    "teacher": teacher, "class_workbook": teacher,
                    "teacher_workbook": "找不到同名教師",
                    "message": f"班別表的 {teacher} 無法對應教師表",
                })
    if empty_teacher:
        warnings.append(f"有 {empty_teacher} 個班別課堂沒有可辨識的教師姓名")
        for lesson in empty_lessons:
            issues.append({
                "code": "missing_teacher", "severity": "error",
                "weekday": lesson["weekday"], "period": lesson["period"],
                "class_code": lesson["class_code"], "subject": lesson["subject"],
                "teacher": "", "class_workbook": "無可辨識教師",
                "teacher_workbook": "—",
                "message": "班別課堂沒有可辨識的教師",
            })

    exact_teacher_keys = {
        (slot["teacher"], slot["weekday"], slot["period"], slot["class_code"])
        for slot in teacher_slots
    }
    slots_by_teacher_time: dict[tuple[str, int, int], list[dict]] = {}
    for slot in teacher_slots:
        slots_by_teacher_time.setdefault(
            (slot["teacher"], slot["weekday"], slot["period"]), []
        ).append(slot)
    for lesson in lessons:
        for teacher in [name for name in lesson["teachers"] if name in known]:
            exact_key = (teacher, lesson["weekday"], lesson["period"], lesson["class_code"])
            if exact_key in exact_teacher_keys:
                continue
            observed = slots_by_teacher_time.get((teacher, lesson["weekday"], lesson["period"]), [])
            teacher_value = "空白／無課" if not observed else "、".join(
                f"{slot['class_code'] or '無班別'} {slot['subject']}".strip() for slot in observed
            )
            issues.append({
                "code": "teacher_slot_mismatch", "severity": "review",
                "resolution_id": f"{lesson['weekday']}:{lesson['period']}:{lesson['class_code']}:{teacher}",
                "weekday": lesson["weekday"], "period": lesson["period"],
                "class_code": lesson["class_code"], "subject": lesson["subject"],
                "teacher": teacher,
                "class_workbook": f"{lesson['class_code']} {lesson['subject']}",
                "teacher_workbook": teacher_value,
                "message": f"{teacher} 在兩份課表的班別／節次不一致",
            })
    lessons_by_slot = {
        (lesson["weekday"], lesson["period"], lesson["class_code"]): lesson
        for lesson in lessons
    }
    for slot in teacher_slots:
        lesson = lessons_by_slot.get((slot["weekday"], slot["period"], slot["class_code"]))
        if not lesson or slot["teacher"] in lesson["teachers"]:
            continue
        issues.append({
            "code": "teacher_extra_assignment", "severity": "review",
            "resolution_id": f"{slot['weekday']}:{slot['period']}:{slot['class_code']}:{slot['teacher']}",
            "weekday": slot["weekday"], "period": slot["period"],
            "class_code": slot["class_code"], "subject": lesson["subject"],
            "teacher": slot["teacher"],
            "class_workbook": f"{lesson['class_code']} {lesson['subject']}（未列此教師）",
            "teacher_workbook": f"{slot['class_code']} {slot['subject']}（列有此教師）",
            "message": f"教師表另列 {slot['teacher']} 教授此課，可加入為共同教師",
        })
    review_count = sum(issue["severity"] == "review" for issue in issues)
    if review_count:
        warnings.append(f"班別表與教師表有 {review_count} 個課堂安排需要人工核對")
    blocked_count = 0
    for lesson in lessons:
        lesson["movable"] = bool(lesson["teachers"]) and all(
            (teacher, lesson["weekday"], lesson["period"], lesson["class_code"]) in exact_teacher_keys
            for teacher in lesson["teachers"]
        )
        if not lesson["movable"]:
            blocked_count += 1
    if blocked_count:
        warnings.append(f"有 {blocked_count} 堂課的班別／教師資料不一致，已鎖定為不可自動調動")

    return {
        "format": "post_exam" if post_exam else "normal",
        "lessons": lessons,
        "teacher_slots": teacher_slots,
        "teachers": sorted(known | {name for lesson in lessons for name in lesson["teachers"]}),
        "summary": {
            "classes": len({lesson["class_code"] for lesson in lessons}),
            "teachers": len(teacher_names),
            "lessons": len(lessons),
            "teacher_slots": len(teacher_slots),
            "blocked_lessons": blocked_count,
            "issues": len(issues),
        },
        "warnings": warnings,
        "issues": issues,
    }


def apply_import_resolutions(payload: dict, resolutions: dict[str, str]) -> dict:
    review_issues = [issue for issue in payload["issues"] if issue["severity"] == "review"]
    for issue in review_issues:
        issue.setdefault(
            "resolution_id",
            f"{issue['weekday']}:{issue['period']}:{issue['class_code']}:{issue['teacher']}",
        )
    expected = {issue["resolution_id"] for issue in review_issues}
    if set(resolutions) != expected or any(value not in {"class", "teacher"} for value in resolutions.values()):
        raise ValueError("請完成所有人工核對項目")

    lessons = payload["lessons"]
    teacher_slots = payload["teacher_slots"]
    chosen_class_slots: dict[tuple[str, int, int], str] = {}
    for issue in review_issues:
        choice = resolutions[issue["resolution_id"]]
        lesson = next(row for row in lessons if (
            row["weekday"], row["period"], row["class_code"], row["subject"]
        ) == (issue["weekday"], issue["period"], issue["class_code"], issue["subject"]))
        if issue["code"] == "teacher_extra_assignment":
            if choice == "teacher":
                if issue["teacher"] not in lesson["teachers"]:
                    lesson["teachers"].append(issue["teacher"])
            else:
                teacher_slots[:] = [slot for slot in teacher_slots if not (
                    slot["teacher"] == issue["teacher"]
                    and slot["weekday"] == issue["weekday"]
                    and slot["period"] == issue["period"]
                    and slot["class_code"] == issue["class_code"]
                )]
            issue["resolution"] = choice
            continue
        teacher_time = (issue["teacher"], issue["weekday"], issue["period"])
        if choice == "class":
            previous_class = chosen_class_slots.get(teacher_time)
            if previous_class and previous_class != issue["class_code"]:
                raise ValueError(f"{issue['teacher']} 同一節不可同時採用兩個班別表安排")
            chosen_class_slots[teacher_time] = issue["class_code"]
            teacher_slots[:] = [slot for slot in teacher_slots if (
                slot["teacher"], slot["weekday"], slot["period"]
            ) != teacher_time]
            teacher_slots.append({
                "teacher": issue["teacher"], "weekday": issue["weekday"], "period": issue["period"],
                "class_code": issue["class_code"], "subject": issue["subject"],
            })
        else:
            lesson["teachers"] = [teacher for teacher in lesson["teachers"] if teacher != issue["teacher"]]
        issue["resolution"] = choice

    exact_teacher_keys = {
        (slot["teacher"], slot["weekday"], slot["period"], slot["class_code"])
        for slot in teacher_slots
    }
    for lesson in lessons:
        lesson["movable"] = bool(lesson["teachers"]) and all(
            (teacher, lesson["weekday"], lesson["period"], lesson["class_code"]) in exact_teacher_keys
            for teacher in lesson["teachers"]
        )
    payload["summary"]["blocked_lessons"] = sum(not lesson["movable"] for lesson in lessons)
    return payload


def get_schedule_revision(db: Session) -> int:
    row = db.get(Configuracio, "schedule_revision")
    try:
        return int(row.valor) if row else 0
    except (TypeError, ValueError):
        return 0


def bump_schedule_revision(db: Session) -> int:
    revision = get_schedule_revision(db) + 1
    row = db.get(Configuracio, "schedule_revision")
    if row:
        row.valor = str(revision)
    else:
        db.add(Configuracio(clau="schedule_revision", valor=str(revision), tipus="integer"))
    return revision


def version_ranges(version: TimetableVersion) -> list[tuple[date, date | None]]:
    if version.effective_ranges_json and is_post_exam_version(version):
        return [(date.fromisoformat(row["effective_from"]), date.fromisoformat(row["effective_to"]))
                for row in json.loads(version.effective_ranges_json)]
    return [(version.effective_from, version.effective_to)]


def serialize_ranges(ranges: list[tuple[date, date | None]]) -> list[dict]:
    return [{"effective_from": start.isoformat(), "effective_to": end.isoformat() if end else None}
            for start, end in ranges]


def matching_range_start(ranges: list[tuple[date, date | None]], day: date) -> date | None:
    return max((start for start, end in ranges if start <= day and (end is None or day <= end)),
               default=None)


def _select_version(candidates, day: date) -> TimetableVersion | None:
    eligible = [(start, version.id, version) for version, ranges in candidates
                if (start := matching_range_start(ranges, day)) is not None]
    return max(eligible, key=lambda row: row[:2])[2] if eligible else None


def version_for_date(db: Session, target_date: date) -> TimetableVersion | None:
    versions = (db.query(TimetableVersion)
            .filter(TimetableVersion.effective_from <= target_date,
                    or_(TimetableVersion.effective_to.is_(None), TimetableVersion.effective_to >= target_date))
            .all())
    return _select_version([(version, version_ranges(version)) for version in versions], target_date)


def is_post_exam_version(version: TimetableVersion) -> bool:
    # ponytail: normal class imports are .xls; persist a type column if they later accept .xlsx.
    return version.class_filename.lower().endswith(".xlsx")


def professor_ids_for_version(db: Session, version: TimetableVersion) -> set[int]:
    teacher_ids = {
        int(teacher)
        for row in db.query(TimetableLesson).filter_by(version_id=version.id).all()
        for teacher in json.loads(row.teachers_json or "[]")
    }
    teacher_ids.update(
        row.professor_id
        for row in db.query(TimetableTeacherSlot).filter_by(version_id=version.id).all()
    )
    return teacher_ids


def _lesson_payload(row: TimetableLesson) -> dict:
    return {
        "lesson_id": row.id,
        "version_id": row.version_id,
        "weekday": row.weekday,
        "period": row.period,
        "class_code": row.class_code,
        "subject": row.subject,
        "teachers": json.loads(row.teachers_json or "[]"),
        "movable": row.movable,
        "special": row.special,
    }


def teaching_dates(db: Session, start: date, future_days: int = 5) -> list[date]:
    closures = {row.data for row in db.query(SchoolClosure).all()}
    result = [start]
    cursor = start
    while len(result) < future_days + 1:
        cursor += timedelta(days=1)
        if cursor.weekday() < 5 and cursor not in closures:
            result.append(cursor)
    return result


def _daterange(start: date, end: date):
    cursor = start
    while cursor <= end:
        yield cursor
        cursor += timedelta(days=1)


def effective_occurrences(db: Session, start: date, end: date) -> list[dict]:
    versions = (db.query(TimetableVersion)
                .filter(TimetableVersion.effective_from <= end,
                        or_(TimetableVersion.effective_to.is_(None), TimetableVersion.effective_to >= start))
                .order_by(TimetableVersion.effective_from, TimetableVersion.id).all())
    if not versions:
        return []

    candidates = [(version, version_ranges(version)) for version in versions]

    lessons_by_version: dict[int, list[TimetableLesson]] = {}
    slots_by_version: dict[int, list[TimetableTeacherSlot]] = {}
    closures = {row.data for row in db.query(SchoolClosure).all()}
    occurrences: list[dict] = []
    for day in _daterange(start, end):
        if day.weekday() >= 5 or day in closures:
            continue
        version = _select_version(candidates, day)
        if not version:
            continue
        if version.id not in lessons_by_version:
            lessons_by_version[version.id] = db.query(TimetableLesson).filter_by(version_id=version.id).all()
        lesson_rows = lessons_by_version[version.id]
        for row in lesson_rows:
            if row.weekday != day.weekday():
                continue
            payload = _lesson_payload(row)
            payload.update({
                "occurrence_id": f"{row.id}:{day.isoformat()}",
                "date": day,
                "locked": not row.movable,
                "adjustment_id": None,
                "source": "base" if row.movable else "base_conflict",
            })
            occurrences.append(payload)

    # A teacher-only slot that does not appear in the class workbook is immutable busy time.
    occupied_base = {
        (teacher, occ["date"], occ["period"])
        for occ in occurrences for teacher in occ["teachers"]
    }
    for day in _daterange(start, end):
        if day.weekday() >= 5 or day in closures:
            continue
        version = _select_version(candidates, day)
        if not version:
            continue
        if version.id not in slots_by_version:
            slots_by_version[version.id] = db.query(TimetableTeacherSlot).filter_by(version_id=version.id).all()
        slots = slots_by_version[version.id]
        for slot in slots:
            if slot.weekday != day.weekday() or (slot.professor_id, day, slot.period) in occupied_base:
                continue
            occurrences.append({
                "occurrence_id": f"slot-{slot.id}:{day.isoformat()}",
                "lesson_id": None,
                "version_id": slot.version_id,
                "weekday": slot.weekday,
                "date": day,
                "period": slot.period,
                "class_code": slot.class_code or "",
                "subject": slot.subject or "",
                "teachers": [slot.professor_id],
                "special": False,
                "locked": True,
                "adjustment_id": None,
                "source": "teacher_workbook",
            })

    confirmed = (db.query(ScheduleAdjustmentLeg, ScheduleAdjustment)
                 .join(ScheduleAdjustment, ScheduleAdjustmentLeg.adjustment_id == ScheduleAdjustment.id)
                 .filter(ScheduleAdjustment.status == "confirmed")
                 .order_by(ScheduleAdjustment.confirmed_at, ScheduleAdjustment.id, ScheduleAdjustmentLeg.id)
                 .all())
    for leg, adjustment in confirmed:
        if adjustment.kind == "co_teacher_solo":
            for occ in occurrences:
                if (occ["lesson_id"] == leg.lesson_id and occ["date"] == leg.from_date
                        and occ["period"] == leg.from_period):
                    occ["teachers"] = [t for t in occ["teachers"] if t != leg.replaced_teacher_id]
                    occ["locked"] = True
                    occ["adjustment_id"] = adjustment.id
                    occ["source"] = adjustment.kind
            continue
        if leg.replacement_teacher_id:
            for occ in occurrences:
                if (occ["lesson_id"] == leg.lesson_id and occ["date"] == leg.from_date
                        and occ["period"] == leg.from_period):
                    teachers = [t for t in occ["teachers"] if t != leg.replaced_teacher_id]
                    teachers.append(leg.replacement_teacher_id)
                    occ["teachers"] = list(dict.fromkeys(teachers))
                    occ["locked"] = True
                    occ["adjustment_id"] = adjustment.id
                    occ["source"] = "emergency_cover"
            continue

        occurrences = [
            occ for occ in occurrences
            if not (occ["lesson_id"] == leg.lesson_id and occ["date"] == leg.from_date
                    and occ["period"] == leg.from_period)
        ]
        if start <= leg.to_date <= end:
            row = db.get(TimetableLesson, leg.lesson_id)
            if not row:
                continue
            occurrences.append({
                "occurrence_id": f"moved-{leg.id}",
                "lesson_id": row.id,
                "version_id": row.version_id,
                "weekday": leg.to_date.weekday(),
                "date": leg.to_date,
                "period": leg.to_period,
                "class_code": leg.class_code,
                "subject": leg.subject,
                "teachers": json.loads(leg.teachers_json or "[]"),
                "special": row.special,
                "locked": True,
                "adjustment_id": adjustment.id,
                "source": adjustment.kind,
            })
    return occurrences


def absence_keys(db: Session, start: date, end: date) -> set[tuple[int, date, int]]:
    rows = (db.query(AbsenceCase)
            .filter(AbsenceCase.status.in_(("open", "resolved")),
                    AbsenceCase.data >= start, AbsenceCase.data <= end).all())
    keys = {
        (row.professor_id, row.data, int(period))
        for row in rows for period in json.loads(row.periods_json or "[]")
    }
    professor_ids = {row.nom: row.id for row in db.query(Professor).all()}
    leaves = (db.query(ProfessorBaixa)
              .filter(ProfessorBaixa.data_inici <= end,
                      ProfessorBaixa.data_final >= start).all())
    for leave in leaves:
        professor_id = professor_ids.get(leave.professor)
        if not professor_id:
            continue
        leave_start = max(start, leave.data_inici)
        leave_end = min(end, leave.data_final)
        for day in _daterange(leave_start, leave_end):
            keys.update((professor_id, day, period) for period in range(1, 10))
    return keys


def _leg(occurrence: dict, to_date: date, to_period: int) -> dict:
    return {
        "occurrence_id": occurrence["occurrence_id"],
        "lesson_id": occurrence["lesson_id"],
        "class_code": occurrence["class_code"],
        "subject": occurrence["subject"],
        "teachers": occurrence["teachers"],
        "special": occurrence.get("special", False),
        "from_date": occurrence["date"].isoformat(),
        "from_period": occurrence["period"],
        "to_date": to_date.isoformat(),
        "to_period": to_period,
    }


def get_max_cycle_lessons(db: Session) -> int:
    row = db.get(Configuracio, MAX_CYCLE_LESSONS_KEY)
    try:
        return min(5, max(2, int(row.valor))) if row else 3
    except (TypeError, ValueError):
        return 3


def _occupancy_index(occurrences: list[dict]) -> tuple[dict[str, dict], dict[tuple, set[str]], dict[tuple, set[str]]]:
    by_id = {occurrence["occurrence_id"]: occurrence for occurrence in occurrences}
    teachers: dict[tuple, set[str]] = {}
    classes: dict[tuple, set[str]] = {}
    for occurrence in occurrences:
        occurrence_id = occurrence["occurrence_id"]
        slot = occurrence["date"], int(occurrence["period"])
        classes.setdefault((occurrence["class_code"], *slot), set()).add(occurrence_id)
        for teacher in occurrence["teachers"]:
            teachers.setdefault((int(teacher), *slot), set()).add(occurrence_id)
    return by_id, teachers, classes


def schedule_slot_started(day: date, period: int, now: datetime | None,
                          period_starts: dict[int, str] | None) -> bool:
    if not now or not period_starts:
        return False
    if day != now.date():
        return day < now.date()
    hour, minute = map(int, period_starts[int(period)].split(":"))
    return (now.hour, now.minute) >= (hour, minute)


def validate_move_legs(legs: list[dict], occurrences: list[dict], absences: set,
                       closures: set[date] | None = None, max_legs: int = 3,
                       occupancy: tuple[dict[str, dict], dict[tuple, set[str]], dict[tuple, set[str]]] | None = None,
                       allowed_locked_ids: set[str] | None = None,
                       now: datetime | None = None,
                       period_starts: dict[int, str] | None = None,
                       ) -> tuple[bool, str]:
    if not 2 <= len(legs) <= max_legs:
        return False, f"調課必須包含 2 至 {max_legs} 堂課"
    classes = {leg["class_code"] for leg in legs}
    if len(classes) != 1:
        return False, "連鎖調課只可在同一班別內進行"
    source_ids = {leg["occurrence_id"] for leg in legs}
    source_map, occupied_teachers, occupied_classes = occupancy or _occupancy_index(occurrences)
    if len(source_ids) != len(legs) or any(source_id not in source_map for source_id in source_ids):
        return False, "課堂來源已改變，請重新分析"
    if len({source_map[source_id]["version_id"] for source_id in source_ids}) != 1:
        return False, "不同課表版本的課堂不能互調"
    allowed_locked_ids = allowed_locked_ids or set()
    if any(source_map[source_id]["locked"] and source_id not in allowed_locked_ids for source_id in source_ids):
        return False, "已確認的課堂已鎖定；請先撤銷原調動"
    if any(schedule_slot_started(source_map[source_id]["date"], source_map[source_id]["period"],
                                 now, period_starts) for source_id in source_ids):
        return False, "來源課堂已開始或已過去"

    teacher_destinations: set[tuple[int, date, int]] = set()
    class_destinations: set[tuple[str, date, int]] = set()
    for leg in legs:
        destination_date = date.fromisoformat(leg["to_date"])
        destination_period = int(leg["to_period"])
        if schedule_slot_started(destination_date, destination_period, now, period_starts):
            return False, "目的節次已開始或已過去"
        if destination_date.weekday() >= 5 or (closures and destination_date in closures):
            return False, "不能把課堂移到非上課日"
        class_key = (leg["class_code"], destination_date, destination_period)
        if class_key in class_destinations:
            return False, "同一班別同一節出現重複課堂"
        class_destinations.add(class_key)
        for teacher in leg["teachers"]:
            key = (int(teacher), destination_date, destination_period)
            if key in teacher_destinations:
                return False, "教師在目的節次重複上課"
            teacher_destinations.add(key)
            if key in absences:
                return False, "教師在目的節次缺席"
            occupants = occupied_teachers.get(key)
            if occupants and not occupants.issubset(source_ids):
                return False, "教師在目的節次已有課"
        occupants = occupied_classes.get(class_key)
        if occupants and not occupants.issubset(source_ids):
            return False, "班別在目的節次已有課"
    return True, ""


def _candidate(kind: str, target: dict, legs: list[dict], start: date, reason: str) -> dict:
    completion = max(date.fromisoformat(leg["to_date"]) for leg in legs)
    special_cross_day_moves = sum(
        bool(leg.get("special")) and leg["from_date"] != leg["to_date"] for leg in legs
    )
    raw = json.dumps({"kind": kind, "legs": legs}, ensure_ascii=False, sort_keys=True)
    return {
        "id": hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16],
        "kind": kind,
        "target_occurrence_id": target["occurrence_id"],
        "completion_date": completion.isoformat(),
        "day_distance": (completion - start).days,
        "moved_lessons": len(legs),
        "special_cross_day_moves": special_cross_day_moves,
        "breaks_consecutive_lessons": 0,
        "new_consecutive_classes": 0,
        "reason": reason,
        "legs": legs,
    }


def _resources(candidate: dict) -> set[tuple]:
    resources: set[tuple] = set()
    for leg in candidate["legs"]:
        resources.add(("occ", leg["occurrence_id"]))
        resources.add(("class", leg["class_code"], leg["to_date"], int(leg["to_period"])))
        for teacher in leg["teachers"]:
            resources.add(("teacher", int(teacher), leg["to_date"], int(leg["to_period"])))
        if leg.get("replacement_teacher_id"):
            resources.add(("teacher", int(leg["replacement_teacher_id"]), leg["to_date"], int(leg["to_period"])))
    return resources


def choose_global(option_groups: list[list[dict]]) -> dict[int, dict]:
    """Maximize solved lessons, then finish sooner and move fewer lessons."""
    order = sorted(range(len(option_groups)), key=lambda index: (len(option_groups[index]) or 999, index))
    resources = {
        id(candidate): _resources(candidate)
        for options in option_groups for candidate in options
    }
    used: set[tuple] = set()
    best: dict[int, dict] = {}
    for index in order:
        for candidate in option_groups[index]:
            candidate_resources = resources[id(candidate)]
            if candidate_resources.isdisjoint(used):
                best[index] = candidate
                used |= candidate_resources
                break

    def score(chosen: dict[int, dict]) -> tuple[int, int, int, int, int, int]:
        return (
            len(chosen),
            -sum(candidate.get("special_cross_day_moves", 0) for candidate in chosen.values()),
            -sum(candidate.get("breaks_consecutive_lessons", 0) for candidate in chosen.values()),
            -sum(candidate.get("new_consecutive_classes", 0) for candidate in chosen.values()),
            -sum(candidate["day_distance"] for candidate in chosen.values()),
            -sum(candidate["moved_lessons"] for candidate in chosen.values()),
        )

    best_score = score(best)
    visits = 0
    # ponytail: bounded exact search; raise the ceiling or use CP-SAT only if real batches outgrow it.
    max_visits = 50_000

    def visit(position: int, used: set[tuple], chosen: dict[int, dict], chosen_score: tuple[int, ...]):
        nonlocal best_score, best, visits
        if visits >= max_visits:
            return
        visits += 1
        if chosen_score[0] + len(order) - position < best_score[0]:
            return
        if position == len(order):
            if chosen_score > best_score:
                best_score, best = chosen_score, dict(chosen)
            return
        index = order[position]
        for candidate in option_groups[index]:
            candidate_resources = resources[id(candidate)]
            if candidate_resources.isdisjoint(used):
                chosen[index] = candidate
                visit(position + 1, used | candidate_resources, chosen, (
                    chosen_score[0] + 1,
                    chosen_score[1] - candidate.get("special_cross_day_moves", 0),
                    chosen_score[2] - candidate.get("breaks_consecutive_lessons", 0),
                    chosen_score[3] - candidate.get("new_consecutive_classes", 0),
                    chosen_score[4] - candidate["day_distance"],
                    chosen_score[5] - candidate["moved_lessons"],
                ))
                chosen.pop(index, None)
        visit(position + 1, used, chosen, chosen_score)

    visit(0, set(), {}, (0, 0, 0, 0, 0, 0))
    return best


def analyze_absences(db: Session, absence_cases: list[AbsenceCase], *,
                     now: datetime | None = None,
                     period_starts: dict[int, str] | None = None) -> dict:
    """Analyse one date's absence cases together so recommendations cannot conflict."""
    if not absence_cases:
        raise ValueError("至少需要一個缺席個案")
    start = absence_cases[0].data
    if any(absence.data != start for absence in absence_cases):
        raise ValueError("同一批缺席必須屬於同一日期")

    dates = teaching_dates(db, start, 5)
    occurrences = effective_occurrences(db, start, dates[-1])
    occurrence_slots = {
        occurrence["occurrence_id"]: (occurrence["date"], int(occurrence["period"]))
        for occurrence in occurrences
    }
    occupancy = _occupancy_index(occurrences)
    consecutive_pairs = _core_consecutive_pairs(occurrences)
    absences = absence_keys(db, start, dates[-1])
    closures = {row.data for row in db.query(SchoolClosure).all()}
    max_cycle_lessons = get_max_cycle_lessons(db)
    targets = []
    for absence in absence_cases:
        periods = {int(value) for value in json.loads(absence.periods_json or "[]")}
        for occurrence in occurrences:
            if (occurrence["date"] == absence.data and occurrence["period"] in periods
                    and absence.professor_id in occurrence["teachers"]
                    and occurrence["lesson_id"] is not None):
                targets.append({
                    **occurrence,
                    "_absence_case_id": absence.id,
                    "_absent_professor_id": absence.professor_id,
                })
    targets.sort(key=lambda occurrence: (occurrence["period"], occurrence["_absence_case_id"]))

    subjects_by_teacher: dict[int, set[str]] = {}
    classes_by_teacher: dict[int, set[str]] = {}
    version = version_for_date(db, start)
    teacher_ids = professor_ids_for_version(db, version) if version else set()
    all_professors = {
        row.id: row.nom for row in db.query(Professor).filter(Professor.id.in_(teacher_ids)).all()
    } if teacher_ids else {}
    if version:
        for row in db.query(TimetableLesson).filter_by(version_id=version.id).all():
            for teacher in json.loads(row.teachers_json or "[]"):
                subjects_by_teacher.setdefault(int(teacher), set()).add(normalize_subject(row.subject))
                classes_by_teacher.setdefault(int(teacher), set()).add(row.class_code)

    option_groups: list[list[dict]] = []
    blocking_reasons: list[str | None] = []
    for target in targets:
        absent_professor_id = target["_absent_professor_id"]
        if any(
            teacher_id != absent_professor_id
            and (teacher_id, target["date"], target["period"]) not in absences
            for teacher_id in target["teachers"]
        ):
            option_groups.append([])
            blocking_reasons.append(None)
            continue
        if schedule_slot_started(target["date"], target["period"], now, period_starts):
            option_groups.append([])
            blocking_reasons.append("started")
            continue
        repairable_target = (
            not target["locked"]
            or bool(target.get("adjustment_id")) and target.get("source") in REPAIRABLE_SWAP_KINDS
        )
        allowed_locked_ids = {target["occurrence_id"]} if repairable_target else set()
        same_class = [
            occ for occ in occurrences
            if occ["class_code"] == target["class_code"] and occ["lesson_id"] is not None
            and occ["occurrence_id"] != target["occurrence_id"] and not occ["locked"]
            and occ["date"] in dates
        ]
        same_class.sort(key=lambda occ: (dates.index(occ["date"]), occ["period"], occ["lesson_id"]))
        candidates: list[dict] = []

        # Priority 1: same-day direct; then each next teaching day, direct before cycle.
        for day in dates:
            direct_for_day: list[dict] = []
            for other in [occ for occ in same_class if occ["date"] == day]:
                legs = [
                    _leg(target, other["date"], other["period"]),
                    _leg(other, target["date"], target["period"]),
                ]
                ok, _ = validate_move_legs(
                    legs, occurrences, absences, closures, occupancy=occupancy,
                    allowed_locked_ids=allowed_locked_ids, now=now, period_starts=period_starts,
                )
                if ok:
                    label = "同日直接互調" if day == start else f"與 {day.isoformat()} 直接互調"
                    candidate = _candidate("direct_swap", target, legs, start, label)
                    candidate["breaks_consecutive_lessons"] = _broken_core_consecutive_count(
                        legs, occurrence_slots, consecutive_pairs
                    )
                    direct_for_day.append(candidate)
            candidates.extend(direct_for_day)

            cycle_for_day: list[dict] = []
            pool = [occ for occ in same_class if occ["date"] <= day][:24]
            for cycle_length in range(3, max_cycle_lessons + 1):
                # ponytail: bounded permutation search; use a solver if chains above five lessons become necessary.
                length_start = len(cycle_for_day)
                for visit, chain in enumerate(permutations(pool, cycle_length - 1), 1):
                    if visit > 10_000 or (cycle_length > 3 and len(cycle_for_day) - length_start >= 200):
                        break
                    if max(occurrence["date"] for occurrence in chain) != day:
                        continue
                    cycle = (target, *chain)
                    legs = [
                        _leg(source, destination["date"], destination["period"])
                        for source, destination in zip(cycle, cycle[1:] + cycle[:1])
                    ]
                    ok, _ = validate_move_legs(
                        legs, occurrences, absences, closures, max_cycle_lessons, occupancy,
                        allowed_locked_ids, now, period_starts,
                    )
                    if ok:
                        label = (f"同班 {cycle_length} 堂連鎖互調" if day == start else
                                 f"最遲於 {day.isoformat()} 完成 {cycle_length} 堂連鎖")
                        candidate = _candidate("three_cycle", target, legs, start, label)
                        candidate["breaks_consecutive_lessons"] = _broken_core_consecutive_count(
                            legs, occurrence_slots, consecutive_pairs
                        )
                        cycle_for_day.append(candidate)
            candidates.extend(cycle_for_day)

        candidates.sort(key=lambda item: (
            bool(item["special_cross_day_moves"]), bool(item["breaks_consecutive_lessons"]),
            item["day_distance"],
            0 if item["kind"] == "direct_swap" else 1,
            item["moved_lessons"], item["id"]
        ))

        # Emergency only when no valid swap exists: same subject, but not a teacher of this class.
        if not candidates and repairable_target:
            for teacher_id, teacher_name in sorted(all_professors.items(), key=lambda item: item[1]):
                if teacher_id == absent_professor_id:
                    continue
                if normalize_subject(target["subject"]) not in subjects_by_teacher.get(teacher_id, set()):
                    continue
                if target["class_code"] in classes_by_teacher.get(teacher_id, set()):
                    continue
                key = (teacher_id, target["date"], target["period"])
                if key in absences:
                    continue
                busy = any(
                    occ["occurrence_id"] != target["occurrence_id"]
                    and occ["date"] == target["date"] and occ["period"] == target["period"]
                    and teacher_id in occ["teachers"]
                    for occ in occurrences
                )
                if busy:
                    continue
                leg = _leg(target, target["date"], target["period"])
                leg["replaced_teacher_id"] = absent_professor_id
                leg["replacement_teacher_id"] = teacher_id
                leg["replacement_teacher_name"] = teacher_name
                candidate = _candidate(
                    "emergency_cover", target, [leg], start,
                    f"迫不得已：由同科的 {teacher_name} 原節代課"
                )
                candidate["new_consecutive_classes"] = adjacent_teaching_count(
                    occurrences, teacher_id, target["date"], int(target["period"])
                )
                candidates.append(candidate)
            candidates.sort(key=lambda item: (
                item["new_consecutive_classes"],
                item["legs"][0]["replacement_teacher_name"],
            ))

        option_groups.append(candidates[:30])
        blocking_reasons.append(None)

    selected = choose_global(option_groups)
    tasks = []
    for index, target in enumerate(targets):
        recommended = selected.get(index)
        alternatives = option_groups[index][:5]
        if recommended and all(item["id"] != recommended["id"] for item in alternatives):
            alternatives = [recommended, *alternatives[:4]]
        tasks.append({
            "task_key": f"{target['_absence_case_id']}:{target['occurrence_id']}",
            "absence_case_id": target["_absence_case_id"],
            "target": serialize_occurrence(target, all_professors),
            "recommended": serialize_candidate(recommended, all_professors),
            "alternatives": [serialize_candidate(candidate, all_professors) for candidate in alternatives],
            "status": "recommended" if index in selected else "unresolved",
            "blocking_reason": blocking_reasons[index],
        })
    return {
        "absence_case_id": absence_cases[0].id if len(absence_cases) == 1 else None,
        "absence_case_ids": [absence.id for absence in absence_cases],
        "revision": get_schedule_revision(db),
        "search_dates": [day.isoformat() for day in dates],
        "tasks": tasks,
        "resolved_count": len(selected),
        "unresolved_count": len(targets) - len(selected),
    }


def analyze_absence(db: Session, absence: AbsenceCase, *, now: datetime | None = None,
                    period_starts: dict[int, str] | None = None) -> dict:
    return analyze_absences(db, [absence], now=now, period_starts=period_starts)


def serialize_occurrence(occurrence: dict, professor_names: dict[int, str]) -> dict:
    return {
        **{key: value for key, value in occurrence.items() if key != "date" and not key.startswith("_")},
        "date": occurrence["date"].isoformat(),
        "teacher_names": [professor_names.get(int(value), str(value)) for value in occurrence["teachers"]],
    }


def serialize_candidate(candidate: dict | None, professor_names: dict[int, str]) -> dict | None:
    if not candidate:
        return None
    result = {**candidate, "legs": []}
    for leg in candidate["legs"]:
        result["legs"].append({
            **leg,
            "teacher_names": [professor_names.get(int(value), str(value)) for value in leg["teachers"]],
        })
    return result


def candidate_from_analysis(
    analysis: dict,
    candidate_id: str,
    absence_case_id: int | None = None,
) -> dict | None:
    for task in analysis["tasks"]:
        if absence_case_id is not None and task.get("absence_case_id") != absence_case_id:
            continue
        candidates = [task.get("recommended"), *task.get("alternatives", [])]
        for candidate in candidates:
            if candidate and candidate["id"] == candidate_id:
                return candidate
    return None
